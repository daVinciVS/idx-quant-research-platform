from __future__ import annotations

import json
import logging
import os
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Any
from zoneinfo import ZoneInfo

import mplfinance as mpf
import numpy as np
import pandas as pd
import requests
import yfinance as yf
from dotenv import load_dotenv
from openpyxl import Workbook
from openpyxl.chart import LineChart, Reference
from openpyxl.drawing.image import Image as OpenpyxlImage
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from pdf_reporter import generate_pdf_report
from src.analytics.trade_plan import calculate_trade_plan
from src.data.contracts import DataContractError, validate_ohlcv
from src.data.market_sessions import exclude_incomplete_daily_dataframe

# ============================================================
# PROJECT CONFIGURATION
# ============================================================

PROJECT_ROOT = Path(__file__).resolve().parent
OUTPUT_DIR = PROJECT_ROOT / "output"
DATA_DIR = PROJECT_ROOT / "data"
LOG_DIR = PROJECT_ROOT / "logs"
CACHE_DIR = DATA_DIR / "cache"

OUTPUT_DIR.mkdir(exist_ok=True)
DATA_DIR.mkdir(exist_ok=True)
LOG_DIR.mkdir(exist_ok=True)
CACHE_DIR.mkdir(exist_ok=True)

LOG_FILE = LOG_DIR / "stock_analysis.log"

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s | %(levelname)s | %(message)s",
    handlers=[
        logging.FileHandler(LOG_FILE, encoding="utf-8"),
        logging.StreamHandler(),
    ],
)

logger = logging.getLogger("idx_stock_analysis")

# ============================================================
# OPTIONAL INDEX ALPHA API CONFIGURATION
# ============================================================

load_dotenv(PROJECT_ROOT / ".env")

INDEX_ALPHA_API_KEY = os.getenv(
    "INDEX_ALPHA_API_KEY",
    "",
).strip()

INDEX_ALPHA_API_BASE_URL = os.getenv(
    "INDEX_ALPHA_API_BASE_URL",
    "https://api.indexalpha.id",
).rstrip("/")

# ============================================================
# EXCEL THEME
# ============================================================

NAVY = "1B365D"
BLUE = "4472C4"
ICE_BLUE = "E8EEF5"
ZEBRA_BLUE = "F4F7FA"

GREEN = "C6EFCE"
GREEN_TEXT = "006100"

RED = "FFC7CE"
RED_TEXT = "9C0006"

YELLOW = "FFEB9C"
YELLOW_TEXT = "9C6500"

ORANGE = "FCE4D6"
WHITE = "FFFFFF"
GRAY = "666666"


# ============================================================
# DATA MODEL
# ============================================================

@dataclass
class YahooStockData:
    ticker: str
    daily_price: pd.DataFrame
    benchmark_price: pd.DataFrame
    info: dict[str, Any]
    balance_sheet: pd.DataFrame
    income_statement: pd.DataFrame
    cash_flow: pd.DataFrame


# ============================================================
# DATA FETCHER
# ============================================================

class StockDataFetcher:

    def __init__(self) -> None:
        self._ihsg_cache: dict[str, pd.DataFrame] = {}

    @staticmethod
    def normalize_ticker(ticker: str) -> str:
        ticker = ticker.strip().upper()

        if not ticker:
            raise ValueError("Ticker code cannot be empty.")

        if not ticker.endswith(".JK"):
            ticker = f"{ticker}.JK"

        return ticker

    @staticmethod
    def clean_history(history: pd.DataFrame) -> pd.DataFrame:
        if history.empty:
            return pd.DataFrame()

        history = history.reset_index()

        dates = pd.to_datetime(
            history["Date"],
            errors="coerce",
        )

        # IDX daily candles must retain the Jakarta trading-date
        # calendar. Converting local midnight to UTC would shift
        # the date back by one day.
        if dates.dt.tz is not None:
            dates = (
                dates
                .dt.tz_convert("Asia/Jakarta")
                .dt.tz_localize(None)
            )

        history["Date"] = dates.dt.normalize()

        return history

    @staticmethod
    def keep_completed_daily_bars(
        history: pd.DataFrame,
    ) -> pd.DataFrame:
        if history.empty:
            return history.copy()

        as_of = datetime.now(tz=ZoneInfo("Asia/Jakarta"))

        completed_history = exclude_incomplete_daily_dataframe(
            history,
            as_of=as_of,
            date_column="Date",
        )

        if len(completed_history) < len(history):
            logger.info(
                "Excluded incomplete current-day candle: %s",
                as_of.strftime("%Y-%m-%d"),
            )

        return completed_history

    @staticmethod
    def validate_daily_ohlcv(
        history: pd.DataFrame,
        *,
        source_name: str,
    ) -> pd.DataFrame:
        """Remove incomplete OHLCV rows, then enforce the data contract."""
        required_columns = [
            "Date",
            "Open",
            "High",
            "Low",
            "Close",
            "Volume",
        ]

        missing_columns = [
            column
            for column in required_columns
            if column not in history.columns
        ]

        if missing_columns:
            raise RuntimeError(
                f"{source_name} response missing columns: "
                f"{missing_columns}"
            )

        cleaned = history.copy()

        for column in required_columns[1:]:
            cleaned[column] = pd.to_numeric(
                cleaned[column],
                errors="coerce",
            )

        before_count = len(cleaned)

        cleaned = cleaned.dropna(subset=required_columns).copy()

        dropped_rows = before_count - len(cleaned)

        if dropped_rows:
            logger.warning(
                "%s: removed %s incomplete OHLCV row(s) "
                "before contract validation.",
                source_name,
                dropped_rows,
            )

        if cleaned.empty:
            raise RuntimeError(
                f"{source_name} has no valid OHLCV rows "
                "after cleaning."
            )

        try:
            return validate_ohlcv(cleaned)
        except DataContractError as error:
            raise RuntimeError(
                f"{source_name} failed OHLCV data validation: {error}"
            ) from error

    def fetch_yahoo_data(
        self,
        ticker: str,
        period: str = "2y",
    ) -> YahooStockData:
        symbol = self.normalize_ticker(ticker)

        logger.info("Fetching Yahoo Finance data for %s", symbol)

        stock = yf.Ticker(symbol)

        daily_price = stock.history(
            period=period,
            interval="1d",
            auto_adjust=False,
            actions=False,
        )

        if daily_price.empty:
            raise RuntimeError(
                f"No historical price data returned for {symbol}. "
                "Check that the ticker is valid."
            )

        daily_price = self.clean_history(daily_price)

        required_columns = [
            "Date",
            "Open",
            "High",
            "Low",
            "Close",
            "Volume",
        ]

        missing_columns = [
            column
            for column in required_columns
            if column not in daily_price.columns
        ]

        if missing_columns:
            raise RuntimeError(
                f"Yahoo response missing: {missing_columns}"
            )

        if "Adj Close" not in daily_price.columns:
            daily_price["Adj Close"] = daily_price["Close"]

        daily_price = daily_price[
            [
                "Date",
                "Open",
                "High",
                "Low",
                "Close",
                "Adj Close",
                "Volume",
            ]
        ].copy()

        daily_price = self.validate_daily_ohlcv(
            daily_price,
            source_name=symbol,
        )

        daily_price["Adj Close"] = pd.to_numeric(
            daily_price["Adj Close"],
            errors="coerce",
        )

        daily_price = daily_price.dropna(
            subset=["Adj Close"]
        ).copy()

        daily_price = self.keep_completed_daily_bars(
            daily_price
        )

        if daily_price.empty:
            raise RuntimeError(
                "No completed daily candles are available "
                "for analysis."
            )

        benchmark_price = self.fetch_ihsg_data(period)

        logger.info(
            "Retrieved %s daily records for %s",
            len(daily_price),
            symbol,
        )

        return YahooStockData(
            ticker=symbol,
            daily_price=daily_price,
            benchmark_price=benchmark_price,
            info=stock.info or {},
            balance_sheet=(
                stock.balance_sheet
                if stock.balance_sheet is not None
                else pd.DataFrame()
            ),
            income_statement=(
                stock.income_stmt
                if stock.income_stmt is not None
                else pd.DataFrame()
            ),
            cash_flow=(
                stock.cashflow
                if stock.cashflow is not None
                else pd.DataFrame()
            ),
        )

    def fetch_ihsg_data(
        self,
        period: str,
    ) -> pd.DataFrame:
        try:
            if period in self._ihsg_cache:
                logger.info(
                    "Using cached IHSG benchmark data"
                )

                return self._ihsg_cache[
                    period
                ].copy()

            logger.info("Fetching IHSG benchmark data")

            ihsg = yf.Ticker("^JKSE")

            benchmark = ihsg.history(
                period=period,
                interval="1d",
                auto_adjust=False,
                actions=False,
            )

            benchmark = self.clean_history(benchmark)
            benchmark = self.validate_daily_ohlcv(
                benchmark,
                source_name="IHSG benchmark",
            )

            if benchmark.empty:
                return pd.DataFrame(
                    columns=["Date", "IHSG Close"]
                )

            benchmark = benchmark[
                ["Date", "Close"]
            ].copy()

            benchmark.rename(
                columns={"Close": "IHSG Close"},
                inplace=True,
            )

            benchmark = self.keep_completed_daily_bars(
                benchmark
            )

            self._ihsg_cache[period] = benchmark.copy()

            return benchmark.copy()

        except Exception as error:
            logger.warning(
                "IHSG benchmark unavailable: %s",
                error,
            )

            return pd.DataFrame(
                columns=["Date", "IHSG Close"]
            )

    def load_local_idx_data(
        self,
        ticker: str,
    ) -> pd.DataFrame:
        """
        Optional future local IDX input.

        If you later create:
        data/local_idx_MDIA.json

        The system will attempt to use it.

        Required fields:
        Date
        Net Foreign Flow
        Top 3 Accumulating Brokers
        Top 3 Distributing Brokers
        Institutional Net Volume
        Retail Net Volume
        Top 3 Net Buy Volume
        Top 1 Buy Volume
        """

        ticker_code = ticker.replace(".JK", "")

        local_file = (
            DATA_DIR
            / f"local_idx_{ticker_code}.json"
        )

        columns = [
            "Date",
            "Net Foreign Flow",
            "Top 3 Accumulating Brokers",
            "Top 3 Distributing Brokers",
            "Institutional Net Volume",
            "Retail Net Volume",
            "Top 3 Net Buy Volume",
            "Top 1 Buy Volume",
        ]

        if not local_file.exists():
            logger.warning(
                "Local IDX data unavailable. "
                "Using N/A Bandarmology placeholders."
            )

            return pd.DataFrame(columns=columns)

        try:
            with local_file.open(
                "r",
                encoding="utf-8",
            ) as handle:
                raw_json = json.load(handle)

            records = raw_json.get(
                "data",
                raw_json,
            )

            local_df = pd.DataFrame(records)

            if "Date" not in local_df.columns:
                raise ValueError(
                    "Local IDX JSON must contain a "
                    "'Date' field in each data record."
                )

            local_df["Date"] = pd.to_datetime(
                local_df["Date"]
            ).dt.normalize()

            for column in columns:
                if column not in local_df.columns:
                    local_df[column] = np.nan

            logger.info(
                "Loaded %s local IDX rows",
                len(local_df),
            )

            return local_df[columns]

        except Exception as error:
            logger.warning(
                "Local IDX data loading failed: %s",
                error,
            )

            return pd.DataFrame(columns=columns)


# ============================================================
# INDEX ALPHA BROKER SUMMARY FETCHER
# ============================================================

class IndexAlphaBrokerFetcher:
    """
    Thin adapter for Index Alpha's /stocks/broker-summary endpoint.

    It works in *aggregated-window* mode, not daily mode.

    For a given ticker and date window (for example, the last 14 days)
    this class will produce one aggregated row per broker code with:

    - net_volume
    - net_value
    - buy_value
    - sell_value

    It then transforms that into a single normalized Bandarmology
    record that matches our AnalyticsEngine expectations:

    Date
    Net Foreign Flow
    Top 3 Accumulating Brokers
    Top 3 Distributing Brokers
    Institutional Net Volume
    Retail Net Volume
    Top 3 Net Buy Volume
    Top 1 Buy Volume

    NOTE:
    - The data is aggregated over the lookback window.
    - We do NOT pretend it is daily data.
    """

    def __init__(self) -> None:
        self.api_key = INDEX_ALPHA_API_KEY
        self.base_url = INDEX_ALPHA_API_BASE_URL

        self.session = requests.Session()
        self.session.headers.update(
            {
                "Accept": "application/json",
                "User-Agent": (
                    "IDXSwingAnalytics/1.0 "
                    "(educational use; contact developer)"
                ),
            }
        )

    def is_configured(self) -> bool:
        placeholder_keys = {
            "",
            "XXXXXXXXXXXXX",
            "YOUR_REAL_KEY_HERE",
        }

        return self.api_key not in placeholder_keys

    @staticmethod
    def _cache_file_name(
        ticker: str,
        from_date: str,
        to_date: str,
        investor: str,
    ) -> Path:
        safe_ticker = ticker.replace(".", "")

        safe_from_date = from_date.replace("-", "")
        safe_to_date = to_date.replace("-", "")

        return CACHE_DIR / (
            f"{safe_ticker}_"
            f"{safe_from_date}_"
            f"{safe_to_date}_"
            f"{investor}.json"
        )

    def _load_cache(
        self,
        ticker: str,
        from_date: str,
        to_date: str,
        investor: str,
    ) -> dict | None:
        cache_file = self._cache_file_name(
            ticker,
            from_date,
            to_date,
            investor,
        )

        if not cache_file.exists():
            return None

        try:
            with cache_file.open(
                "r",
                encoding="utf-8",
            ) as handle:
                logger.info(
                    "Using cached Index Alpha data: %s",
                    cache_file.name,
                )

                return json.load(handle)

        except Exception as error:
            logger.warning(
                "Failed to read cache file %s (%s).",
                cache_file.name,
                error,
            )

            return None

    def _save_cache(
        self,
        ticker: str,
        from_date: str,
        to_date: str,
        investor: str,
        payload: dict,
    ) -> None:
        cache_file = self._cache_file_name(
            ticker,
            from_date,
            to_date,
            investor,
        )

        try:
            with cache_file.open(
                "w",
                encoding="utf-8",
            ) as handle:
                json.dump(
                    payload,
                    handle,
                    ensure_ascii=False,
                    indent=2,
                )

            logger.info(
                "Saved Index Alpha API cache: %s",
                cache_file.name,
            )

        except Exception as error:
            logger.warning(
                "Failed to write cache file %s (%s).",
                cache_file.name,
                error,
            )

    def _compute_date_window(
        self,
        df_dates: pd.Series,
        lookback_days: int,
    ) -> tuple[str, str, pd.Timestamp]:
        """
        Use the actual available market dates in df_dates.

        from_date = last_date - lookback_days + 1
        to_date   = last_date
        """

        last_date = pd.to_datetime(df_dates.max()).normalize()
        from_date = last_date - pd.Timedelta(
            days=lookback_days - 1
        )

        return (
            from_date.strftime("%Y-%m-%d"),
            last_date.strftime("%Y-%m-%d"),
            last_date,
        )

    def _call_broker_summary(
        self,
        ticker: str,
        from_date: str,
        to_date: str,
        investor: str,
    ) -> dict:
        """
        Low-level HTTP call to GET /stocks/broker-summary.

        investor can be:
        - "all"
        - "foreign"
        - "domestic"
        """

        url = (
            f"{self.base_url}"
            "/stocks/broker-summary"
        )

        headers = {
            "Authorization": (
                f"Bearer {self.api_key}"
            ),
        }

        params = {
            "ticker": ticker,
            "from": from_date,
            "to": to_date,
            "investor": investor,
            "market": "RG",  # Regular market only
        }

        response = self.session.get(
            url,
            headers=headers,
            params=params,
            timeout=20,
        )

        if response.status_code == 401:
            raise RuntimeError(
                "Index Alpha authentication failed. "
                "Check INDEX_ALPHA_API_KEY in your .env file."
            )

        if not response.ok:
            raise RuntimeError(
                f"Index Alpha API error "
                f"({response.status_code}): "
                f"{response.text[:200]}"
            )

        return response.json()

    @staticmethod
    def _aggregate_brokers(
        payload: dict,
    ) -> pd.DataFrame:
        """
        Convert Index Alpha BrokerSummaryResponse into
        a per-broker aggregated table.

        Expected payload structure (simplified):

        {
          "data": [
            {
              "broker": "AK",
              "buy_volume": ...,
              "sell_volume": ...,
              "buy_value": ...,
              "sell_value": ...
            },
            ...
          ]
        }
        """

        data = payload.get("data") or []

        rows = []

        for item in data:
            broker = str(
                item.get("code")
                or item.get("broker")
                or ""
            ).upper()

            if not broker:
                continue

            buy_volume = float(
                item.get(
                    "buy_volume",
                    0,
                )
                or 0
            )

            sell_volume = float(
                item.get(
                    "sell_volume",
                    0,
                )
                or 0
            )

            buy_value = float(
                item.get(
                    "buy_value",
                    0,
                )
                or 0
            )

            sell_value = float(
                item.get(
                    "sell_value",
                    0,
                )
                or 0
            )

            net_volume = (
                buy_volume - sell_volume
            )

            net_value = (
                buy_value - sell_value
            )

            rows.append(
                {
                    "broker": broker,
                    "buy_volume": buy_volume,
                    "sell_volume": sell_volume,
                    "buy_value": buy_value,
                    "sell_value": sell_value,
                    "net_volume": net_volume,
                    "net_value": net_value,
                }
            )

        if not rows:
            return pd.DataFrame(
                columns=[
                    "broker",
                    "buy_volume",
                    "sell_volume",
                    "buy_value",
                    "sell_value",
                    "net_volume",
                    "net_value",
                ]
            )

        df = pd.DataFrame(rows)

        # If multiple entries per broker exist, aggregate them.
        return df.groupby("broker", as_index=False).sum()

    def fetch_aggregated_window(
        self,
        ticker: str,
        price_dates: pd.Series,
        lookback_days: int = 14,
        use_cache: bool = True,
    ) -> tuple[pd.DataFrame, pd.Timestamp, int]:
        """
        High-level method to obtain an aggregated 14-day (or N-day)
        broker-summary window for:

        - investor="all"   -> overall broker behavior
        - investor="foreign" -> foreign-investor broker behavior

        Returns:
        - normalized Bandarmology row as a DataFrame (0 or 1 row)
        - last_date corresponding to the window end
        - api_calls_used (0, 1, or 2)
        """

        if not self.is_configured():
            logger.info(
                "Index Alpha API key not configured. "
                "Skipping broker-summary integration."
            )
            return (
                pd.DataFrame(
                    columns=[
                        "Date",
                        "Net Foreign Flow",
                        "Top 3 Accumulating Brokers",
                        "Top 3 Distributing Brokers",
                        "Institutional Net Volume",
                        "Retail Net Volume",
                        "Top 3 Net Buy Volume",
                        "Top 1 Buy Volume",
                    ]
                ),
                None,
                0,
            )

        from_date, to_date, last_date = (
            self._compute_date_window(
                price_dates,
                lookback_days,
            )
        )

        logger.info(
            "Broker summary window for %s: %s to %s "
            "(%s days)",
            ticker,
            from_date,
            to_date,
            lookback_days,
        )

        api_calls_used = 0

        # --- Investor = ALL ---
        all_payload = None

        if use_cache:
            all_payload = self._load_cache(
                ticker,
                from_date,
                to_date,
                "all",
            )

        if all_payload is None:
            all_payload = self._call_broker_summary(
                ticker,
                from_date,
                to_date,
                investor="all",
            )
            self._save_cache(
                ticker,
                from_date,
                to_date,
                "all",
                all_payload,
            )
            api_calls_used += 1

        all_df = self._aggregate_brokers(
            all_payload
        )

        # --- Investor = FOREIGN ---
        foreign_payload = None

        if use_cache:
            foreign_payload = self._load_cache(
                ticker,
                from_date,
                to_date,
                "foreign",
            )

        if foreign_payload is None:
            foreign_payload = (
                self._call_broker_summary(
                    ticker,
                    from_date,
                    to_date,
                    investor="f",
                )
            )
            self._save_cache(
                ticker,
                from_date,
                to_date,
                "foreign",
                foreign_payload,
            )
            api_calls_used += 1

        foreign_df = self._aggregate_brokers(
            foreign_payload
        )

        if all_df.empty:
            logger.warning(
                "Index Alpha returned no broker summary data "
                "for %s in window %s to %s.",
                ticker,
                from_date,
                to_date,
            )

            return (
                pd.DataFrame(
                    columns=[
                        "Date",
                        "Net Foreign Flow",
                        "Top 3 Accumulating Brokers",
                        "Top 3 Distributing Brokers",
                        "Institutional Net Volume",
                        "Retail Net Volume",
                        "Top 3 Net Buy Volume",
                        "Top 1 Buy Volume",
                    ]
                ),
                last_date,
                api_calls_used,
            )

        # ----------------------------------------------------
        # Normalize aggregated ALL-broker data.
        # ----------------------------------------------------

        sorted_by_net_value = all_df.sort_values(
            "net_value",
            ascending=False,
        )

        top_buyers = (
            sorted_by_net_value.head(3)
        )

        top_sellers = (
            sorted_by_net_value.tail(3)
            .sort_values(
                "net_value",
                ascending=True,
            )
        )

        def format_broker_list(
            broker_df: pd.DataFrame,
        ) -> str:
            formatted_brokers = []

            for _, broker_row in broker_df.iterrows():
                broker_code = broker_row["broker"]
                net_value = broker_row["net_value"]

                value_in_billions = (
                    net_value / 1_000_000_000
                )

                sign = "+" if net_value >= 0 else ""

                formatted_brokers.append(
                    
                        f"{broker_code} "
                        f"({sign}IDR "
                        f"{value_in_billions:,.2f}B)"
                    
                )

            return ", ".join(formatted_brokers)

        top3_buy_codes = format_broker_list(
            top_buyers
        )

        top3_sell_codes = format_broker_list(
            top_sellers
        )

        top3_net_buy_volume = float(
            top_buyers["net_volume"]
            .clip(lower=0)
            .sum()
        )

        top1_buy_volume = float(
            max(
                [
                    max(
                        v,
                        0,
                    )
                    for v in top_buyers[
                        "net_volume"
                    ].tolist()
                ]
                or [0]
            )
        )

        # For now, we use:
        # - foreign_df aggregated net_value as Net Foreign Flow proxy
        # - total net_volume as Institutional vs Retail proxy placeholder
        net_foreign_flow_value = 0.0

        if not foreign_df.empty:
            net_foreign_flow_value = float(
                foreign_df["net_value"].sum()
            )

        institutional_net_volume = float(
            all_df["net_volume"].sum()
        )

        # Retail net volume is not explicitly known
        # from this endpoint; treat it as 0 for now
        # to avoid making strong claims about
        # ownership from broker codes alone.
        retail_net_volume = 0.0

        normalized_row = {
            "Date": last_date,
            "Net Foreign Flow": net_foreign_flow_value,
            "Top 3 Accumulating Brokers": (
                top3_buy_codes
                or "N/A - Data Unavailable"
            ),
            "Top 3 Distributing Brokers": (
                top3_sell_codes
                or "N/A - Data Unavailable"
            ),
            "Institutional Net Volume": institutional_net_volume,
            "Retail Net Volume": retail_net_volume,
            "Top 3 Net Buy Volume": top3_net_buy_volume,
            "Top 1 Buy Volume": top1_buy_volume,
        }

        return (
            pd.DataFrame([normalized_row]),
            last_date,
            api_calls_used,
        )

# ============================================================
# MANUAL MIRAE CSV DATA LOADER
# ============================================================

class ManualMiraeDataLoader:
    """
    Loads manually entered broker summary and foreign flow data.

    Expected files:

    data/manual/broker_summary_manual.csv
    data/manual/foreign_flow_manual.csv
    """

    def __init__(self) -> None:
        self.manual_dir = DATA_DIR / "manual"

        self.broker_file = (
            self.manual_dir
            / "broker_summary_manual.csv"
        )

        self.foreign_file = (
            self.manual_dir
            / "foreign_flow_manual.csv"
        )

    @staticmethod
    def empty_dataframe() -> pd.DataFrame:
        return pd.DataFrame(
            columns=[
                "Date",
                "Net Foreign Flow",
                "Top 3 Accumulating Brokers",
                "Top 3 Distributing Brokers",
                "Institutional Net Volume",
                "Retail Net Volume",
                "Top 3 Net Buy Volume",
                "Top 1 Buy Volume",
            ]
        )

    @staticmethod
    def format_broker_list(
        broker_df: pd.DataFrame,
    ) -> str:
        formatted = []

        for _, broker_row in broker_df.iterrows():
            broker_code = str(
                broker_row["Broker_Code"]
            ).upper()

            net_volume = broker_row["Net_Volume"]
            net_value = broker_row["Net_Value"]

            if pd.notna(net_value):
                value_billions = (
                    net_value / 1_000_000_000
                )

                sign = "+" if net_value >= 0 else ""

                formatted.append(
                    
                        f"{broker_code} "
                        f"({sign}IDR "
                        f"{value_billions:,.2f}B)"
                    
                )
            else:
                sign = "+" if net_volume >= 0 else ""

                formatted.append(
                    
                        f"{broker_code} "
                        f"({sign}"
                        f"{net_volume:,.0f} sh)"
                    
                )

        return ", ".join(formatted)

    def load(
        self,
        ticker: str,
    ) -> pd.DataFrame:
        ticker_code = (
            ticker
            .replace(".JK", "")
            .strip()
            .upper()
        )

        broker_daily = pd.DataFrame()
        foreign_daily = pd.DataFrame()

        # ----------------------------------------------------
        # Load manual broker-summary CSV
        # ----------------------------------------------------

        if self.broker_file.exists():
            try:
                broker_df = pd.read_csv(
                    self.broker_file
                )

                required_broker_columns = [
                    "Date",
                    "Ticker",
                    "Broker_Code",
                    "Net_Volume",
                    "Net_Value",
                ]

                missing_columns = [
                    column
                    for column in required_broker_columns
                    if column not in broker_df.columns
                ]

                if missing_columns:
                    raise ValueError(
                        "Missing broker CSV columns: "
                        f"{missing_columns}"
                    )

                broker_df["Date"] = pd.to_datetime(
                    broker_df["Date"]
                ).dt.normalize()

                broker_df["Ticker"] = (
                    broker_df["Ticker"]
                    .astype(str)
                    .str.upper()
                    .str.replace(".JK", "", regex=False)
                )

                broker_df = broker_df[
                    broker_df["Ticker"] == ticker_code
                ].copy()

                broker_df["Net_Volume"] = pd.to_numeric(
                    broker_df["Net_Volume"],
                    errors="coerce",
                ).fillna(0)

                broker_df["Net_Value"] = pd.to_numeric(
                    broker_df["Net_Value"],
                    errors="coerce",
                )

                daily_rows = []

                for date, date_group in broker_df.groupby(
                    "Date"
                ):
                    # If at least one Net_Value is provided,
                    # rank brokers by Net_Value.
                    #
                    # Otherwise rank by Net_Volume.
                    if date_group["Net_Value"].notna().any():
                        ranking_column = "Net_Value"
                    else:
                        ranking_column = "Net_Volume"

                    top_buyers = (
                        date_group[
                            date_group[ranking_column] > 0
                        ]
                        .sort_values(
                            ranking_column,
                            ascending=False,
                        )
                        .head(3)
                    )

                    top_sellers = (
                        date_group[
                            date_group[ranking_column] < 0
                        ]
                        .sort_values(
                            ranking_column,
                            ascending=True,
                        )
                        .head(3)
                    )

                    top3_net_buy_volume = float(
                        top_buyers["Net_Volume"]
                        .clip(lower=0)
                        .sum()
                    )

                    top1_buy_volume = float(
                        top_buyers["Net_Volume"]
                        .clip(lower=0)
                        .max()
                        if not top_buyers.empty
                        else 0
                    )

                    daily_rows.append(
                        {
                            "Date": date,
                            "Top 3 Accumulating Brokers": (
                                self.format_broker_list(
                                    top_buyers
                                )
                                if not top_buyers.empty
                                else (
                                    "N/A - Data Unavailable"
                                )
                            ),
                            "Top 3 Distributing Brokers": (
                                self.format_broker_list(
                                    top_sellers
                                )
                                if not top_sellers.empty
                                else (
                                    "N/A - Data Unavailable"
                                )
                            ),
                            "Top 3 Net Buy Volume": (
                                top3_net_buy_volume
                            ),
                            "Top 1 Buy Volume": (
                                top1_buy_volume
                            ),
                        }
                    )

                broker_daily = pd.DataFrame(
                    daily_rows
                )

                logger.info(
                    "Loaded %s manual broker-summary dates "
                    "for %s.",
                    len(broker_daily),
                    ticker_code,
                )

            except Exception as error:
                logger.warning(
                    "Manual broker-summary CSV failed: %s",
                    error,
                )

        else:
            logger.warning(
                "Manual broker-summary CSV not found: %s",
                self.broker_file,
            )

        # ----------------------------------------------------
        # Load manual foreign-flow CSV
        # ----------------------------------------------------

        if self.foreign_file.exists():
            try:
                foreign_df = pd.read_csv(
                    self.foreign_file
                )

                required_foreign_columns = [
                    "Date",
                    "Ticker",
                    "Foreign_Buy_Value",
                    "Foreign_Sell_Value",
                ]

                missing_columns = [
                    column
                    for column in required_foreign_columns
                    if column not in foreign_df.columns
                ]

                if missing_columns:
                    raise ValueError(
                        "Missing foreign-flow CSV columns: "
                        f"{missing_columns}"
                    )

                foreign_df["Date"] = pd.to_datetime(
                    foreign_df["Date"]
                ).dt.normalize()

                foreign_df["Ticker"] = (
                    foreign_df["Ticker"]
                    .astype(str)
                    .str.upper()
                    .str.replace(".JK", "", regex=False)
                )

                foreign_df = foreign_df[
                    foreign_df["Ticker"] == ticker_code
                ].copy()

                foreign_df["Foreign_Buy_Value"] = (
                    pd.to_numeric(
                        foreign_df["Foreign_Buy_Value"],
                        errors="coerce",
                    ).fillna(0)
                )

                foreign_df["Foreign_Sell_Value"] = (
                    pd.to_numeric(
                        foreign_df["Foreign_Sell_Value"],
                        errors="coerce",
                    ).fillna(0)
                )

                foreign_df["Net Foreign Flow"] = (
                    foreign_df["Foreign_Buy_Value"]
                    - foreign_df["Foreign_Sell_Value"]
                )

                foreign_daily = (
                    foreign_df.groupby(
                        "Date",
                        as_index=False,
                    )["Net Foreign Flow"]
                    .sum()
                )

                logger.info(
                    "Loaded %s manual foreign-flow dates "
                    "for %s.",
                    len(foreign_daily),
                    ticker_code,
                )

            except Exception as error:
                logger.warning(
                    "Manual foreign-flow CSV failed: %s",
                    error,
                )

        else:
            logger.warning(
                "Manual foreign-flow CSV not found: %s",
                self.foreign_file,
            )

        # ----------------------------------------------------
        # Combine broker and foreign-flow daily records.
        # ----------------------------------------------------

        if broker_daily.empty and foreign_daily.empty:
            return self.empty_dataframe()

        if broker_daily.empty:
            combined = foreign_daily.copy()
        elif foreign_daily.empty:
            combined = broker_daily.copy()
        else:
            combined = broker_daily.merge(
                foreign_daily,
                on="Date",
                how="outer",
            )

        for column in [
            "Top 3 Accumulating Brokers",
            "Top 3 Distributing Brokers",
        ]:
            if column not in combined.columns:
                combined[column] = (
                    "N/A - Data Unavailable"
                )

            combined[column] = combined[column].fillna(
                "N/A - Data Unavailable"
            )

        for column in [
            "Net Foreign Flow",
            "Top 3 Net Buy Volume",
            "Top 1 Buy Volume",
        ]:
            if column not in combined.columns:
                combined[column] = np.nan

        # We intentionally do not estimate institutional vs
        # retail ownership from manual broker codes.
        combined["Institutional Net Volume"] = np.nan
        combined["Retail Net Volume"] = np.nan

        output_columns = [
            "Date",
            "Net Foreign Flow",
            "Top 3 Accumulating Brokers",
            "Top 3 Distributing Brokers",
            "Institutional Net Volume",
            "Retail Net Volume",
            "Top 3 Net Buy Volume",
            "Top 1 Buy Volume",
        ]

        return (
            combined[output_columns]
            .sort_values("Date")
            .reset_index(drop=True)
        )

# ============================================================
# ANALYTICS ENGINE
# ============================================================

class AnalyticsEngine:
    def calculate_indicators(
        self,
        daily_price: pd.DataFrame,
        benchmark_price: pd.DataFrame,
        local_idx_data: pd.DataFrame,
    ) -> pd.DataFrame:
        df = daily_price.copy()

        # ----------------------------------------------------
        # Merge IHSG benchmark
        # ----------------------------------------------------

        if not benchmark_price.empty:
            df = df.merge(
                benchmark_price,
                on="Date",
                how="left",
            )
        else:
            df["IHSG Close"] = np.nan

        # ----------------------------------------------------
        # Merge optional local IDX data
        # ----------------------------------------------------

        if not local_idx_data.empty:
            df = df.merge(
                local_idx_data,
                on="Date",
                how="left",
            )
        else:
            df["Net Foreign Flow"] = np.nan
            df["Top 3 Accumulating Brokers"] = (
                "N/A - Data Unavailable"
            )
            df["Top 3 Distributing Brokers"] = (
                "N/A - Data Unavailable"
            )
            df["Institutional Net Volume"] = np.nan
            df["Retail Net Volume"] = np.nan
            df["Top 3 Net Buy Volume"] = np.nan
            df["Top 1 Buy Volume"] = np.nan

        local_columns = [
            "Net Foreign Flow",
            "Top 3 Accumulating Brokers",
            "Top 3 Distributing Brokers",
            "Institutional Net Volume",
            "Retail Net Volume",
            "Top 3 Net Buy Volume",
            "Top 1 Buy Volume",
        ]

        for column in local_columns:
            if column not in df.columns:
                df[column] = np.nan

        df["Top 3 Accumulating Brokers"] = (
            df["Top 3 Accumulating Brokers"]
            .fillna("N/A - Data Unavailable")
        )

        df["Top 3 Distributing Brokers"] = (
            df["Top 3 Distributing Brokers"]
            .fillna("N/A - Data Unavailable")
        )

        # ----------------------------------------------------
        # Returns and volume
        # ----------------------------------------------------

        df["Daily Return"] = df["Close"].pct_change()

        df["Log Return"] = np.log(
            df["Close"] / df["Close"].shift(1)
        )

        df["20D Vol SMA"] = (
            df["Volume"]
            .rolling(window=20, min_periods=20)
            .mean()
        )

        df["Volume Ratio"] = (
            df["Volume"] / df["20D Vol SMA"]
        )

        # ----------------------------------------------------
        # Up-Volume vs Down-Volume Context
        #
        # Green candle:
        # Close > Open
        #
        # Red candle:
        # Close < Open
        #
        # This separates volume participation on bullish
        # days from volume participation on bearish days.
        # ----------------------------------------------------

        df["Green Candle"] = (
            df["Close"] > df["Open"]
        )

        df["Red Candle"] = (
            df["Close"] < df["Open"]
        )

        df["Bull Volume"] = np.where(
            df["Green Candle"],
            df["Volume"],
            0,
        )

        df["Bear Volume"] = np.where(
            df["Red Candle"],
            df["Volume"],
            0,
        )

        df["Bull Volume Ratio"] = np.where(
            df["Green Candle"],
            df["Volume Ratio"],
            np.nan,
        )

        df["Bear Volume Ratio"] = np.where(
            df["Red Candle"],
            df["Volume Ratio"],
            np.nan,
        )

        df["Bull Volume 5D"] = (
            df["Bull Volume"]
            .rolling(window=5, min_periods=1)
            .sum()
        )

        df["Bear Volume 5D"] = (
            df["Bear Volume"]
            .rolling(window=5, min_periods=1)
            .sum()
        )

        df["Total Volume 5D"] = (
            df["Volume"]
            .rolling(window=5, min_periods=1)
            .sum()
        )

        df["Bull Volume Share 5D"] = (
            df["Bull Volume 5D"]
            / df["Total Volume 5D"]
        )

        df["Bear Volume Share 5D"] = (
            df["Bear Volume 5D"]
            / df["Total Volume 5D"]
        )

        df["Bull Volume Ratio 5D"] = (
            df["Bull Volume Ratio"]
            .rolling(window=5, min_periods=1)
            .mean()
        )

        df["Bear Volume Ratio 5D"] = (
            df["Bear Volume Ratio"]
            .rolling(window=5, min_periods=1)
            .mean()
        )


        # ----------------------------------------------------
        # Trend indicators
        # ----------------------------------------------------

        df["SMA20"] = (
            df["Close"]
            .rolling(window=20, min_periods=20)
            .mean()
        )

        df["SMA50"] = (
            df["Close"]
            .rolling(window=50, min_periods=50)
            .mean()
        )

        df["SMA150"] = (
            df["Close"]
            .rolling(window=150, min_periods=150)
            .mean()
        )

        df["SMA200"] = (
            df["Close"]
            .rolling(window=200, min_periods=200)
            .mean()
        )

        # ----------------------------------------------------
        # Final accumulation/distribution flags.
        #
        # These must be calculated after SMA20 exists.
        # ----------------------------------------------------

        df["Institutional Accumulation Flag"] = (
            (
                df["Bull Volume Share 5D"] > 0.55
            )
            & (
                df["Bull Volume Ratio 5D"]
                > (
                    df["Bear Volume Ratio 5D"]
                    * 1.20
                )
            )
            & (
                df["Close"] > df["SMA20"]
            )
        )

        df["Distribution Pressure Flag"] = (
            (
                df["Bear Volume Share 5D"] > 0.55
            )
            & (
                df["Bear Volume Ratio 5D"]
                > (
                    df["Bull Volume Ratio 5D"]
                    * 1.20
                )
            )
            & (
                df["Close"] < df["SMA20"]
            )
        )

        df["EMA9"] = (
            df["Close"]
            .ewm(span=9, adjust=False)
            .mean()
        )

        df["EMA21"] = (
            df["Close"]
            .ewm(span=21, adjust=False)
            .mean()
        )

        # ----------------------------------------------------
        # ATR 14
        # ----------------------------------------------------

        previous_close = df["Close"].shift(1)

        high_low = df["High"] - df["Low"]
        high_previous_close = (
            df["High"] - previous_close
        ).abs()

        low_previous_close = (
            df["Low"] - previous_close
        ).abs()

        df["True Range"] = pd.concat(
            [
                high_low,
                high_previous_close,
                low_previous_close,
            ],
            axis=1,
        ).max(axis=1)

        df["ATR14"] = (
            df["True Range"]
            .ewm(
                alpha=1 / 14,
                adjust=False,
                min_periods=14,
            )
            .mean()
        )

        df["ATR Percent"] = (
            df["ATR14"] / df["Close"]
        )

        # ----------------------------------------------------
        # RSI 14
        # ----------------------------------------------------

        price_change = df["Close"].diff()

        gains = price_change.clip(lower=0)
        losses = -price_change.clip(upper=0)

        average_gain = gains.ewm(
            alpha=1 / 14,
            adjust=False,
            min_periods=14,
        ).mean()

        average_loss = losses.ewm(
            alpha=1 / 14,
            adjust=False,
            min_periods=14,
        ).mean()

        relative_strength = (
            average_gain
            / average_loss.replace(0, np.nan)
        )

        df["RSI14"] = 100 - (
            100 / (1 + relative_strength)
        )

        # ----------------------------------------------------
        # MACD 12, 26, 9
        # ----------------------------------------------------

        ema12 = df["Close"].ewm(
            span=12,
            adjust=False,
        ).mean()

        ema26 = df["Close"].ewm(
            span=26,
            adjust=False,
        ).mean()

        df["MACD"] = ema12 - ema26

        df["MACD Signal"] = (
            df["MACD"]
            .ewm(span=9, adjust=False)
            .mean()
        )

        df["MACD Histogram"] = (
            df["MACD"] - df["MACD Signal"]
        )

        # ----------------------------------------------------
        # Bollinger Bands 20, 2
        # ----------------------------------------------------

        df["BB Middle"] = (
            df["Close"]
            .rolling(window=20, min_periods=20)
            .mean()
        )

        bb_std = (
            df["Close"]
            .rolling(window=20, min_periods=20)
            .std()
        )

        df["BB Upper"] = (
            df["BB Middle"] + (2 * bb_std)
        )

        df["BB Lower"] = (
            df["BB Middle"] - (2 * bb_std)
        )

        df["BB Width"] = (
            (df["BB Upper"] - df["BB Lower"])
            / df["BB Middle"]
        )

        # ----------------------------------------------------
        # OBV
        # ----------------------------------------------------

        price_direction = np.sign(
            df["Close"].diff()
        ).fillna(0)

        df["OBV"] = (
            price_direction * df["Volume"]
        ).cumsum()

        # ----------------------------------------------------
        # ADX 14, +DI, -DI
        # ----------------------------------------------------

        up_move = df["High"].diff()
        down_move = -df["Low"].diff()

        plus_dm = np.where(
            (up_move > down_move) & (up_move > 0),
            up_move,
            0,
        )

        minus_dm = np.where(
            (down_move > up_move) & (down_move > 0),
            down_move,
            0,
        )

        atr_smooth = df["True Range"].ewm(
            alpha=1 / 14,
            adjust=False,
            min_periods=14,
        ).mean()

        plus_dm_smooth = pd.Series(
            plus_dm,
            index=df.index,
        ).ewm(
            alpha=1 / 14,
            adjust=False,
            min_periods=14,
        ).mean()

        minus_dm_smooth = pd.Series(
            minus_dm,
            index=df.index,
        ).ewm(
            alpha=1 / 14,
            adjust=False,
            min_periods=14,
        ).mean()

        df["Plus DI"] = (
            100
            * plus_dm_smooth
            / atr_smooth.replace(0, np.nan)
        )

        df["Minus DI"] = (
            100
            * minus_dm_smooth
            / atr_smooth.replace(0, np.nan)
        )

        dx = (
            100
            * (df["Plus DI"] - df["Minus DI"]).abs()
            / (
                df["Plus DI"] + df["Minus DI"]
            ).replace(0, np.nan)
        )

        df["ADX14"] = dx.ewm(
            alpha=1 / 14,
            adjust=False,
            min_periods=14,
        ).mean()

        # ----------------------------------------------------
        # Support and resistance
        # ----------------------------------------------------

        df["Support 20D"] = (
            df["Low"]
            .rolling(window=20, min_periods=20)
            .min()
        )

        # Previous 20 completed sessions' highest high.
        #
        # shift(1) excludes the current candle, so today's
        # close can be evaluated against a real prior ceiling.
        df["Resistance 20D"] = (
            df["High"]
            .shift(1)
            .rolling(window=20, min_periods=20)
            .max()
        )

        # ----------------------------------------------------
        # Relative strength versus IHSG
        # ----------------------------------------------------

        df["Relative Strength vs IHSG"] = (
            df["Close"] / df["IHSG Close"]
        )

        df["RS vs IHSG 20D Change"] = (
            df["Relative Strength vs IHSG"]
            .pct_change(
                periods=20,
                fill_method=None,
            )
        )

        # ----------------------------------------------------
        # Bandarmology rolling metrics
        # ----------------------------------------------------

        df["14D Foreign Flow"] = (
            df["Net Foreign Flow"]
            .rolling(window=14, min_periods=1)
            .sum()
        )

        df["14D Top 3 Net Buy"] = (
            df["Top 3 Net Buy Volume"]
            .rolling(window=14, min_periods=1)
            .sum()
        )

        logger.info(
            "Indicators calculated for %s records",
            len(df),
        )

        return df

    @staticmethod
    def get_statement_value(
        statement: pd.DataFrame,
        labels: list[str],
    ) -> float | None:
        if statement.empty:
            return None

        for label in labels:
            if label in statement.index:
                values = pd.to_numeric(
                    statement.loc[label],
                    errors="coerce",
                ).dropna()

                if not values.empty:
                    return float(values.iloc[0])

        return None

    def extract_fundamentals(
        self,
        info: dict[str, Any],
        balance_sheet: pd.DataFrame,
        income_statement: pd.DataFrame,
        cash_flow: pd.DataFrame,
    ) -> dict[str, Any]:
        cash = self.get_statement_value(
            balance_sheet,
            [
                "Cash Cash Equivalents And Short Term Investments",
                "Cash And Cash Equivalents",
                "Cash Financial",
            ],
        )

        short_term_debt = self.get_statement_value(
            balance_sheet,
            [
                "Current Debt",
                "Current Debt And Capital Lease Obligation",
            ],
        )

        long_term_debt = self.get_statement_value(
            balance_sheet,
            [
                "Long Term Debt",
                "Long Term Debt And Capital Lease Obligation",
            ],
        )

        total_assets = self.get_statement_value(
            balance_sheet,
            ["Total Assets"],
        )

        total_equity = self.get_statement_value(
            balance_sheet,
            [
                "Stockholders Equity",
                "Total Equity Gross Minority Interest",
            ],
        )

        net_income = self.get_statement_value(
            income_statement,
            [
                "Net Income",
                "Net Income Common Stockholders",
            ],
        )

        operating_cash_flow = self.get_statement_value(
            cash_flow,
            [
                "Operating Cash Flow",
                "Total Cash From Operating Activities",
            ],
        )

        total_debt = (
            (short_term_debt or 0)
            + (long_term_debt or 0)
        )

        debt_to_equity = None

        if total_equity not in [None, 0]:
            debt_to_equity = total_debt / total_equity

        return {
            "ticker": info.get("symbol", "N/A"),
            "name": (
                info.get("longName")
                or info.get("shortName")
                or "N/A"
            ),
            "sector": info.get("sector", "N/A"),
            "industry": info.get("industry", "N/A"),
            "market_cap": info.get("marketCap"),
            "enterprise_value": info.get("enterpriseValue"),
            "trailing_pe": info.get("trailingPE"),
            "forward_pe": info.get("forwardPE"),
            "price_to_book": info.get("priceToBook"),
            "cash": cash,
            "short_term_debt": short_term_debt,
            "long_term_debt": long_term_debt,
            "total_debt": total_debt,
            "total_assets": total_assets,
            "total_equity": total_equity,
            "net_debt": total_debt - (cash or 0),
            "debt_to_equity": debt_to_equity,
            "current_ratio": info.get("currentRatio"),
            "quick_ratio": info.get("quickRatio"),
            "net_income": net_income,
            "operating_cash_flow": operating_cash_flow,
        }

    @staticmethod
    def trend_bias(
        close: float,
        moving_average: float,
    ) -> str:
        if pd.isna(moving_average):
            return "N/A"

        if close > moving_average:
            return "BULLISH"

        if close < moving_average:
            return "BEARISH"

        return "NEUTRAL"

    def evaluate_extension_risk(
        self,
        latest: pd.Series,
    ) -> dict[str, Any]:
        """
        Detects whether price is too extended above SMA20.

        EXTENDED:
        RSI14 > 70
        AND Close > SMA20 × 1.15

        BUYING CLIMAX / SEVERELY OVEREXTENDED:
        RSI14 > 75
        AND Close > SMA20 × 1.25
        AND Volume Ratio > 1.50x
        """

        close = latest["Close"]
        sma20 = latest["SMA20"]
        rsi14 = latest["RSI14"]
        volume_ratio = latest["Volume Ratio"]

        if pd.isna(sma20) or sma20 <= 0:
            return {
                "status": "N/A",
                "distance_from_sma20": np.nan,
                "reason": (
                    "SMA20 is unavailable."
                ),
            }

        distance_from_sma20 = (
            close / sma20
        ) - 1

        buying_climax = (
            pd.notna(rsi14)
            and pd.notna(volume_ratio)
            and rsi14 > 75
            and close > (sma20 * 1.25)
            and volume_ratio > 1.50
        )

        extended = (
            pd.notna(rsi14)
            and rsi14 > 70
            and close > (sma20 * 1.15)
        )

        if buying_climax:
            return {
                "status": (
                    "BUYING CLIMAX / "
                    "SEVERELY OVEREXTENDED"
                ),
                "distance_from_sma20": (
                    distance_from_sma20
                ),
                "reason": (
                    f"RSI14 is {rsi14:.2f}, "
                    f"price is {distance_from_sma20:.1%} "
                    "above SMA20, and volume ratio is "
                    f"{volume_ratio:.2f}x."
                ),
            }

        if extended:
            return {
                "status": "EXTENDED",
                "distance_from_sma20": (
                    distance_from_sma20
                ),
                "reason": (
                    f"RSI14 is {rsi14:.2f} and "
                    f"price is {distance_from_sma20:.1%} "
                    "above SMA20."
                ),
            }

        return {
            "status": "NORMAL",
            "distance_from_sma20": (
                distance_from_sma20
            ),
            "reason": (
                "Price is not materially extended "
                "above SMA20."
            ),
        }

    def evaluate_wyckoff_phase(
        self,
        df: pd.DataFrame,
    ) -> str:
        """
        Wyckoff-inspired market-structure classifier.

        Confirmed accumulation/distribution requires:
        - A tight 30-day consolidation range.
        - Price near SMA50.
        - Non-conflicting broker / foreign-flow confirmation.
        - Supporting bullish or bearish volume behavior.

        Candidate labels can use price-volume evidence when
        broker and foreign-flow data is unavailable, such as
        during Yahoo-only batch screening.
        """

        latest = df.iloc[-1]

        close = latest["Close"]
        sma50 = latest["SMA50"]
        sma200 = latest["SMA200"]

        recent_30_days = df.tail(30)

        foreign_flow = latest["14D Foreign Flow"]
        broker_flow = latest["14D Top 3 Net Buy"]

        accumulation_volume_flag = (
            bool(
                latest[
                    "Institutional Accumulation Flag"
                ]
            )
            if pd.notna(
                latest[
                    "Institutional Accumulation Flag"
                ]
            )
            else False
        )

        distribution_volume_flag = (
            bool(
                latest[
                    "Distribution Pressure Flag"
                ]
            )
            if pd.notna(
                latest[
                    "Distribution Pressure Flag"
                ]
            )
            else False
        )

        # ----------------------------------------------------
        # Long-term trend structure
        # ----------------------------------------------------

        markup_structure = (
            pd.notna(sma50)
            and pd.notna(sma200)
            and close > sma50
            and sma50 > sma200
        )

        markdown_structure = (
            pd.notna(sma50)
            and pd.notna(sma200)
            and close < sma50
            and sma50 < sma200
        )

        # ----------------------------------------------------
        # Tight 30-day consolidation structure
        # ----------------------------------------------------

        recent_high = recent_30_days["High"].max()
        recent_low = recent_30_days["Low"].min()

        range_percent = np.nan

        if recent_low > 0:
            range_percent = (
                (recent_high - recent_low)
                / recent_low
            )

        price_near_sma50 = (
            pd.notna(sma50)
            and sma50 > 0
            and abs(close - sma50) / sma50 <= 0.05
        )

        sideways_structure = (
            pd.notna(range_percent)
            and range_percent <= 0.10
            and price_near_sma50
        )

        # ----------------------------------------------------
        # Flow direction and conflict handling
        # ----------------------------------------------------

        foreign_positive = (
            pd.notna(foreign_flow)
            and foreign_flow > 0
        )

        foreign_negative = (
            pd.notna(foreign_flow)
            and foreign_flow < 0
        )

        broker_positive = (
            pd.notna(broker_flow)
            and broker_flow > 0
        )

        broker_negative = (
            pd.notna(broker_flow)
            and broker_flow < 0
        )

        flow_conflict = (
            (
                foreign_positive
                and broker_negative
            )
            or (
                foreign_negative
                and broker_positive
            )
        )

        flow_positive = (
            not flow_conflict
            and (
                foreign_positive
                or broker_positive
            )
        )

        flow_negative = (
            not flow_conflict
            and (
                foreign_negative
                or broker_negative
            )
        )

        # ----------------------------------------------------
        # Classification priority
        #
        # Confirmed phases require structure, flow, and
        # supportive volume. Candidate phases are less strict.
        # ----------------------------------------------------

        if (
            sideways_structure
            and flow_positive
            and accumulation_volume_flag
        ):
            return "ACCUMULATION PHASE"

        if (
            sideways_structure
            and flow_negative
            and distribution_volume_flag
        ):
            return "DISTRIBUTION PHASE"

        if (
            sideways_structure
            and not flow_conflict
            and (
                flow_positive
                or accumulation_volume_flag
            )
            and not distribution_volume_flag
        ):
            return "ACCUMULATION CANDIDATE"

        if (
            sideways_structure
            and not flow_conflict
            and (
                flow_negative
                or distribution_volume_flag
            )
            and not accumulation_volume_flag
        ):
            return "DISTRIBUTION CANDIDATE"

        if markup_structure:
            return "MARKUP PHASE"

        if markdown_structure:
            return "MARKDOWN PHASE"

        return "TRANSITION / INCONCLUSIVE"

    def evaluate_gorengan_risk(
        self,
        fundamentals: dict[str, Any],
        df: pd.DataFrame,
    ) -> tuple[int, str, list[str]]:
        latest = df.iloc[-1]

        points = 0
        reasons = []

        market_cap = fundamentals["market_cap"]

        if (
            market_cap is not None
            and market_cap < 1_000_000_000_000
        ):
            points += 1
            reasons.append(
                "Market cap is below IDR 1 trillion."
            )

        if (
            pd.notna(latest["ATR Percent"])
            and latest["ATR Percent"] > 0.05
        ):
            points += 1
            reasons.append(
                "ATR volatility exceeds 5% of price."
            )

        if (
            pd.notna(latest["Top 1 Buy Volume"])
            and latest["Volume"] > 0
            and (
                latest["Top 1 Buy Volume"]
                / latest["Volume"]
            ) > 0.70
        ):
            points += 1
            reasons.append(
                "Single broker concentration exceeds 70%."
            )

        average_trading_value = (
            df.tail(20)["Close"]
            * df.tail(20)["Volume"]
        ).mean()

        if average_trading_value < 1_000_000_000:
            points += 1
            reasons.append(
                "Average daily traded value is below IDR 1 billion."
            )

        if points == 0:
            label = "SAFE (Bluechip / Liquid)"
        elif points <= 2:
            label = "MODERATE RISK (Second Liner)"
        else:
            label = "EXTREME RISK (Saham Gorengan)"

        if not reasons:
            reasons.append(
                "No risk conditions were triggered."
            )

        return points, label, reasons

    def calculate_metrics(
        self,
        df: pd.DataFrame,
        fundamentals: dict[str, Any],
        report_df: pd.DataFrame | None = None,
    ) -> dict[str, Any]:
        latest = df.iloc[-1]

        minervini = self.evaluate_minervini_template(
            df
        )

        extension_risk = self.evaluate_extension_risk(
            latest
        )

        reporting_window = (
            report_df.copy()
            if report_df is not None
            and not report_df.empty
            else df.copy()
        )

        technical_score = 0
        fundamental_score = 0
        bandarmology_score = 0
        evidence = []

        close = latest["Close"]

        # ----------------------------------------------------
        # Technical decision score
        # ----------------------------------------------------

        if (
            pd.notna(latest["SMA20"])
            and pd.notna(latest["SMA50"])
            and close > latest["SMA20"]
            and close > latest["SMA50"]
        ):
            technical_score += 11
            evidence.append(
                [
                    "Trend",
                    "Positive",
                    11,
                    (
                        f"Close {close:,.2f} is above "
                        f"SMA20 {latest['SMA20']:,.2f} and "
                        f"SMA50 {latest['SMA50']:,.2f}."
                    ),
                ]
            )
        else:
            evidence.append(
                [
                    "Trend",
                    "Negative / Neutral",
                    0,
                    "Close is not above both SMA20 and SMA50.",
                ]
            )

        bull_volume_share = latest[
            "Bull Volume Share 5D"
        ]

        bear_volume_share = latest[
            "Bear Volume Share 5D"
        ]

        bull_volume_ratio_5d = latest[
            "Bull Volume Ratio 5D"
        ]

        bear_volume_ratio_5d = latest[
            "Bear Volume Ratio 5D"
        ]

        if latest["Institutional Accumulation Flag"]:
            technical_score += 11

            evidence.append(
                [
                    "Up-Volume Context",
                    "Accumulation",
                    11,
                    (
                        f"Bull volume share is "
                        f"{bull_volume_share:.1%} over 5 days. "
                        f"Bull relative volume is "
                        f"{bull_volume_ratio_5d:.2f}x versus "
                        f"bear relative volume at "
                        f"{bear_volume_ratio_5d:.2f}x. "
                        "Price is above SMA20."
                    ),
                ]
            )

        elif latest["Distribution Pressure Flag"]:
            evidence.append(
                [
                    "Up-Volume Context",
                    "Distribution Pressure",
                    0,
                    (
                        f"Bear volume share is "
                        f"{bear_volume_share:.1%} over 5 days. "
                        f"Bear relative volume is "
                        f"{bear_volume_ratio_5d:.2f}x versus "
                        f"bull relative volume at "
                        f"{bull_volume_ratio_5d:.2f}x. "
                        "Price is below SMA20."
                    ),
                ]
            )

        else:
            evidence.append(
                [
                    "Up-Volume Context",
                    "Neutral",
                    0,
                    (
                        "No strong 5-day accumulation or "
                        "distribution-volume imbalance was detected."
                    ),
                ]
            )

        if (
            pd.notna(latest["RSI14"])
            and 45 <= latest["RSI14"] <= 65
        ):
            technical_score += 11
            evidence.append(
                [
                    "RSI Momentum",
                    "Positive",
                    11,
                    (
                        f"RSI14 is {latest['RSI14']:.2f}, "
                        "inside the constructive 45 to 65 zone."
                    ),
                ]
            )
        else:
            rsi_text = (
                f"{latest['RSI14']:.2f}"
                if pd.notna(latest["RSI14"])
                else "N/A"
            )

            evidence.append(
                [
                    "RSI Momentum",
                    "Neutral / Extended",
                    0,
                    (
                        f"RSI14 is {rsi_text}. "
                        "Preferred buy-zone condition was not met."
                    ),
                ]
            )

        # ----------------------------------------------------
        # Fundamental decision score
        # ----------------------------------------------------

        de_ratio = fundamentals["debt_to_equity"]

        if de_ratio is not None and de_ratio < 1.2:
            fundamental_score += 11
            evidence.append(
                [
                    "Debt Structure",
                    "Positive",
                    11,
                    f"Debt-to-equity is {de_ratio:.2f}, below 1.20.",
                ]
            )
        else:
            evidence.append(
                [
                    "Debt Structure",
                    "Neutral / Negative",
                    0,
                    "Debt-to-equity is unavailable or above 1.20.",
                ]
            )

        # ----------------------------------------------------
        # Profitability score:
        # Net Income = maximum 7 points
        # Operating Cash Flow = maximum 4 points
        #
        # Missing data is excluded from the denominator.
        # Known negative data receives 0 points.
        # ----------------------------------------------------

        net_income = fundamentals["net_income"]
        operating_cash_flow = (
            fundamentals["operating_cash_flow"]
        )

        net_income_available = net_income is not None
        operating_cash_flow_available = (
            operating_cash_flow is not None
        )

        profitability_score = 0
        profitability_max_score = 0

        if net_income_available:
            profitability_max_score += 7

            if net_income > 0:
                profitability_score += 7

                evidence.append(
                    [
                        "Net Income",
                        "Positive",
                        7,
                        (
                            "Net income is positive: "
                            f"IDR {net_income:,.0f}."
                        ),
                    ]
                )
            else:
                evidence.append(
                    [
                        "Net Income",
                        "Negative",
                        0,
                        (
                            "Net income is negative: "
                            f"IDR {net_income:,.0f}."
                        ),
                    ]
                )
        else:
            evidence.append(
                [
                    "Net Income",
                    "Not Scored",
                    0,
                    (
                        "Net income is unavailable from Yahoo Finance "
                        "and is excluded from the score denominator."
                    ),
                ]
            )

        if operating_cash_flow_available:
            profitability_max_score += 4

            if operating_cash_flow > 0:
                profitability_score += 4

                evidence.append(
                    [
                        "Operating Cash Flow",
                        "Positive",
                        4,
                        (
                            "Operating cash flow is positive: "
                            f"IDR {operating_cash_flow:,.0f}."
                        ),
                    ]
                )
            else:
                evidence.append(
                    [
                        "Operating Cash Flow",
                        "Negative",
                        0,
                        (
                            "Operating cash flow is negative: "
                            f"IDR {operating_cash_flow:,.0f}."
                        ),
                    ]
                )
        else:
            evidence.append(
                [
                    "Operating Cash Flow",
                    "Not Scored",
                    0,
                    (
                        "Operating cash flow is unavailable from Yahoo Finance "
                        "and is excluded from the score denominator."
                    ),
                ]
            )

        fundamental_score += profitability_score

        pb = fundamentals["price_to_book"]

        if pb is not None and pb < 2.0:
            fundamental_score += 11
            evidence.append(
                [
                    "Valuation",
                    "Positive",
                    11,
                    f"Price-to-book is {pb:.2f}, below 2.00.",
                ]
            )
        else:
            evidence.append(
                [
                    "Valuation",
                    "Neutral",
                    0,
                    "Price-to-book is unavailable or above 2.00.",
                ]
            )

        # ----------------------------------------------------
        # Bandarmology decision score
        # ----------------------------------------------------

        if (
            pd.notna(latest["14D Top 3 Net Buy"])
            and latest["14D Top 3 Net Buy"] > 0
        ):
            bandarmology_score += 17

            top3_net_buy = latest["14D Top 3 Net Buy"]

            broker_summary_text = latest[
                "Top 3 Accumulating Brokers"
            ]

            evidence.append(
                [
                    "Broker Accumulation",
                    "Positive",
                    17,
                    (
                        f"Top-3 net-buy volume is "
                        f"{top3_net_buy:,.0f} shares across "
                        f"the selected aggregation window. "
                        f"Top brokers: {broker_summary_text}"
                    ),
                ]
            )
        else:
            evidence.append(
                [
                    "Broker Accumulation",
                    "Unavailable / Neutral",
                    0,
                    (
                        "Broker-summary data is unavailable "
                        "or Top-3 net-buy volume was not positive."
                    ),
                ]
            )

        if (
            pd.notna(latest["14D Foreign Flow"])
            and latest["14D Foreign Flow"] > 0
        ):
            bandarmology_score += 17

            foreign_flow_value = latest[
                "14D Foreign Flow"
            ]

            foreign_flow_billions = (
                foreign_flow_value
                / 1_000_000_000
            )

            evidence.append(
                [
                    "Foreign Investor Flow",
                    "Positive",
                    17,
                    (
                        "Foreign-investor broker-flow proxy is "
                        f"+IDR {foreign_flow_billions:,.2f}B "
                        "across the selected aggregation window."
                    ),
                ]
            )
        else:
            foreign_flow_text = "N/A"

            if pd.notna(latest["14D Foreign Flow"]):
                foreign_flow_billions = (
                    latest["14D Foreign Flow"]
                    / 1_000_000_000
                )

                foreign_flow_text = (
                    f"IDR {foreign_flow_billions:,.2f}B"
                )

            evidence.append(
                [
                    "Foreign Investor Flow",
                    "Unavailable / Neutral",
                    0,
                    (
                        "Foreign-investor broker-flow proxy is "
                        f"{foreign_flow_text}. "
                        "Positive accumulation was not confirmed."
                    ),
                ]
            )

        # ----------------------------------------------------
        # Dynamic score denominator
        # Missing data is excluded from the maximum score.
        # ----------------------------------------------------

        broker_data_available = (
            df["Top 3 Net Buy Volume"]
            .notna()
            .any()
        )

        foreign_flow_data_available = (
            df["Net Foreign Flow"]
            .notna()
            .any()
        )

        technical_max_score = 33
        fundamental_max_score = (
            22 + profitability_max_score
        )

        broker_max_score = (
            17
            if broker_data_available
            else 0
        )

        foreign_flow_max_score = (
            17
            if foreign_flow_data_available
            else 0
        )

        bandarmology_max_score = (
            broker_max_score
            + foreign_flow_max_score
        )

        eligible_max_score = (
            technical_max_score
            + fundamental_max_score
            + bandarmology_max_score
        )

        raw_score = (
            technical_score
            + fundamental_score
            + bandarmology_score
        )

        normalized_score = 0.0

        if eligible_max_score > 0:
            normalized_score = (
                raw_score
                / eligible_max_score
                * 100
            )

        data_coverage_percent = (
            eligible_max_score
            / 100
            * 100
        )

        total_score = raw_score
        if minervini["passed"]:
            evidence.append(
                [
                    "Minervini Trend Template",
                    "Passed",
                    0,
                    (
                        "All 4 structural trend checks passed: "
                        "Close > SMA150, Close > SMA200, "
                        "SMA50 > SMA150, and price is at least "
                        "30% above the 52-week low."
                    ),
                ]
            )
        else:
            failed_rules_text = " ".join(
                minervini["failed_rules"]
            )

            evidence.append(
                [
                    "Minervini Trend Template",
                    "Failed",
                    0,
                    failed_rules_text,
                ]
            )

        if (
            extension_risk["status"]
            == "BUYING CLIMAX / SEVERELY OVEREXTENDED"
        ):
            evidence.append(
                [
                    "Extension Risk",
                    "Buying Climax",
                    0,
                    (
                        "WITHHOLD NEW ENTRY. "
                        + extension_risk["reason"]
                    ),
                ]
            )

        elif extension_risk["status"] == "EXTENDED":
            evidence.append(
                [
                    "Extension Risk",
                    "Extended",
                    0,
                    (
                        "Avoid chasing a stretched price. "
                        + extension_risk["reason"]
                    ),
                ]
            )

        else:
            evidence.append(
                [
                    "Extension Risk",
                    "Normal",
                    0,
                    extension_risk["reason"],
                ]
            )

        # ----------------------------------------------------
        # Trade plan calculations
        # ----------------------------------------------------

        atr14 = latest["ATR14"]

        support = latest["Support 20D"]
        resistance = latest["Resistance 20D"]

        six_month_high = reporting_window["High"].max()
        six_month_low = reporting_window["Low"].min()

        # ----------------------------------------------------
        # Two independent swing-trading scenarios:
        #
        # 1. Pullback Plan:
        #    Enter near current price after a controlled pullback.
        #
        # 2. Breakout Plan:
        #    Enter only after a daily close breaks above 20D resistance.
        #
        # Each plan has its own stop-loss, target, and RRR.
        # ----------------------------------------------------

        pullback_entry_low = np.nan
        pullback_entry_high = close
        pullback_stop_loss = np.nan
        pullback_target_1 = np.nan
        pullback_target_2 = np.nan
        pullback_rrr = np.nan

        breakout_entry = np.nan
        breakout_stop_loss = np.nan
        breakout_target_1 = np.nan
        breakout_target_2 = np.nan
        breakout_rrr = np.nan

        if (
            pd.notna(atr14)
            and pd.notna(resistance)
            and pd.notna(six_month_high)
        ):
            try:
                trade_plan = calculate_trade_plan(
                    close=float(close),
                    atr14=float(atr14),
                    resistance=float(resistance),
                    six_month_high=float(six_month_high),
                )
            except (TypeError, ValueError) as error:
                logger.warning(
                    "Trade plan unavailable due to invalid inputs: %s",
                    error,
                )
            else:
                pullback_entry_low = trade_plan.pullback_entry_low
                pullback_entry_high = trade_plan.pullback_entry_high
                pullback_stop_loss = trade_plan.pullback_stop_loss
                pullback_target_1 = trade_plan.pullback_target_1
                pullback_target_2 = trade_plan.pullback_target_2
                pullback_rrr = trade_plan.pullback_rrr

                breakout_entry = trade_plan.breakout_entry
                breakout_stop_loss = trade_plan.breakout_stop_loss
                breakout_target_1 = trade_plan.breakout_target_1
                breakout_target_2 = trade_plan.breakout_target_2
                breakout_rrr = trade_plan.breakout_rrr

        # Keep these aliases temporarily so older dashboard
        # components continue to work before we redesign Tab 2.
        entry_low = pullback_entry_low
        entry_high = pullback_entry_high
        stop_loss = pullback_stop_loss
        target_1 = pullback_target_1
        target_2 = pullback_target_2
        reward_risk_ratio = pullback_rrr

        risk_points, risk_label, risk_reasons = (
            self.evaluate_gorengan_risk(
                fundamentals,
                df,
            )
        )

        # ----------------------------------------------------
        # Final decision
        # ----------------------------------------------------

        # ----------------------------------------------------
        # Decision uses normalized score, not raw score.
        # ----------------------------------------------------

        bandarmology_complete = (
            broker_data_available
            and foreign_flow_data_available
        )

        if normalized_score >= 75:
            if bandarmology_complete:
                decision = "STRONG BUY / INVEST"
                confidence = "HIGH"
            else:
                decision = "BUY CANDIDATE - FLOW UNVERIFIED"
                confidence = "MEDIUM"

        elif normalized_score >= 55:
            decision = "HOLD / WATCHLIST"
            confidence = "MEDIUM"

        else:
            decision = "AVOID / DO NOT INVEST"
            confidence = "LOW"

        if (
            pd.notna(reward_risk_ratio)
            and reward_risk_ratio < 2.0
            and decision in [
                "STRONG BUY / INVEST",
                "BUY CANDIDATE - FLOW UNVERIFIED",
            ]
        ):
            decision = "HOLD / WATCHLIST"
            confidence = "MEDIUM"

            evidence.append(
                [
                    "Reward-to-Risk",
                    "Downgraded",
                    0,
                    (
                        f"RRR is {reward_risk_ratio:.2f}, "
                        "below the preferred minimum of 2.00."
                    ),
                ]
            )

        # ----------------------------------------------------
        # Minervini Trend Template gate:
        # A stock in a weak long-term trend cannot receive
        # a trade-ready bullish decision.
        # ----------------------------------------------------

        if not minervini["passed"]:
            decision = "AVOID / TREND TEMPLATE FAILED"
            confidence = "LOW"

        # ----------------------------------------------------
        # Extension-risk decision gate:
        #
        # A strong trend can still be too stretched to buy.
        # ----------------------------------------------------

        if (
            extension_risk["status"]
            == "BUYING CLIMAX / SEVERELY OVEREXTENDED"
            and not decision.startswith("AVOID")
        ):
            decision = (
                "WITHHOLD NEW ENTRY / TAKE PROFITS"
            )

            confidence = "LOW"

        elif (
            extension_risk["status"] == "EXTENDED"
            and decision in [
                "STRONG BUY / INVEST",
                "BUY CANDIDATE - FLOW UNVERIFIED",
            ]
        ):
            decision = "HOLD / WAIT FOR PULLBACK"

            confidence = "MEDIUM"

        institutional_net = df.tail(14)[
            "Institutional Net Volume"
        ].sum(min_count=1)

        retail_net = df.tail(14)[
            "Retail Net Volume"
        ].sum(min_count=1)

        smart_money_percent = np.nan
        retail_percent = np.nan

        if (
            pd.notna(institutional_net)
            and pd.notna(retail_net)
        ):
            total_flow = (
                abs(institutional_net)
                + abs(retail_net)
            )

            if total_flow > 0:
                smart_money_percent = (
                    abs(institutional_net)
                    / total_flow
                    * 100
                )

                retail_percent = 100 - smart_money_percent

        # ----------------------------------------------------
        # Extra indicator interpretations
        # ----------------------------------------------------

        if (
            pd.notna(latest["EMA9"])
            and pd.notna(latest["EMA21"])
            and latest["EMA9"] > latest["EMA21"]
        ):
            ema_signal = "BULLISH"

        else:
            ema_signal = "BEARISH / NEUTRAL"

        if (
            pd.notna(latest["MACD"])
            and pd.notna(latest["MACD Signal"])
            and latest["MACD"] > latest["MACD Signal"]
        ):
            macd_signal = "BULLISH MOMENTUM"

        else:
            macd_signal = "WEAKENING / BEARISH MOMENTUM"

        if pd.notna(latest["ADX14"]):
            if latest["ADX14"] >= 25:
                adx_signal = "STRONG TREND"
            elif latest["ADX14"] >= 20:
                adx_signal = "MODERATE TREND"
            else:
                adx_signal = "WEAK / RANGING"
        else:
            adx_signal = "N/A"

        if (
            pd.notna(latest["RS vs IHSG 20D Change"])
            and latest["RS vs IHSG 20D Change"] > 0
        ):
            relative_strength_signal = (
                "OUTPERFORMING IHSG"
            )
        else:
            relative_strength_signal = (
                "UNDERPERFORMING / N/A"
            )

        return {
            "decision": decision,
            "minervini_passed": minervini["passed"],
            "minervini_passed_checks": (
                minervini["passed_checks"]
            ),
            "minervini_total_checks": (
                minervini["total_checks"]
            ),
            "minervini_sma150": minervini["sma150"],
            "minervini_sma200": minervini["sma200"],
            "minervini_52w_low": (
                minervini["fifty_two_week_low"]
            ),
            "minervini_failed_rules": (
                minervini["failed_rules"]
            ),
            "extension_risk_status": (
                extension_risk["status"]
            ),
            "extension_distance_from_sma20": (
                extension_risk["distance_from_sma20"]
            ),
            "extension_risk_reason": (
                extension_risk["reason"]
            ),
            "raw_score": raw_score,
            "eligible_max_score": eligible_max_score,
            "normalized_score": normalized_score,
            "data_coverage_percent": data_coverage_percent,
            "broker_data_available": broker_data_available,
            "foreign_flow_data_available": foreign_flow_data_available,
            "bandarmology_complete": bandarmology_complete,
            "confidence": confidence,
            "total_score": total_score,
            "technical_score": technical_score,
            "fundamental_score": fundamental_score,
            "bandarmology_score": bandarmology_score,
            "evidence": evidence,
            "wyckoff_phase": self.evaluate_wyckoff_phase(df),
            "risk_points": risk_points,
            "risk_label": risk_label,
            "risk_reasons": risk_reasons,
            "latest_date": latest["Date"],
            "latest_close": close,
            "latest_volume": latest["Volume"],
            "sma20": latest["SMA20"],
            "sma50": latest["SMA50"],
            "ema9": latest["EMA9"],
            "ema21": latest["EMA21"],
            "atr14": atr14,
            "atr_percent": latest["ATR Percent"],
            "rsi14": latest["RSI14"],
            "volume_ratio": latest["Volume Ratio"],
                        "bull_volume_share_5d": (
                latest["Bull Volume Share 5D"]
            ),
            "bear_volume_share_5d": (
                latest["Bear Volume Share 5D"]
            ),
            "bull_volume_ratio_5d": (
                latest["Bull Volume Ratio 5D"]
            ),
            "bear_volume_ratio_5d": (
                latest["Bear Volume Ratio 5D"]
            ),
            "institutional_accumulation_flag": (
                latest["Institutional Accumulation Flag"]
            ),
            "distribution_pressure_flag": (
                latest["Distribution Pressure Flag"]
            ),
            "macd": latest["MACD"],
            "macd_signal_line": latest["MACD Signal"],
            "macd_histogram": latest["MACD Histogram"],
            "adx14": latest["ADX14"],
            "plus_di": latest["Plus DI"],
            "minus_di": latest["Minus DI"],
            "bb_upper": latest["BB Upper"],
            "bb_middle": latest["BB Middle"],
            "bb_lower": latest["BB Lower"],
            "support": support,
            "resistance": resistance,
            "trend_sma20": self.trend_bias(
                close,
                latest["SMA20"],
            ),
            "trend_sma50": self.trend_bias(
                close,
                latest["SMA50"],
            ),
            "ema_signal": ema_signal,
            "macd_signal": macd_signal,
            "adx_signal": adx_signal,
            "relative_strength_signal": (
                relative_strength_signal
            ),
            "daily_volatility": (
                reporting_window["Daily Return"].std()
            ),
            "annualized_volatility": (
                reporting_window["Daily Return"].std()
                * np.sqrt(252)
            ),
            "six_month_high": six_month_high,
            "six_month_low": six_month_low,
            "maximum_drawdown": (
                reporting_window["Close"]
                / reporting_window["Close"].cummax()
                - 1
            ).min(),
            "entry_low": entry_low,
            "entry_high": entry_high,
            "breakout_entry": breakout_entry,
            "stop_loss": stop_loss,
            "target_1": target_1,
            "target_2": target_2,
            "reward_risk_ratio": reward_risk_ratio,

            "pullback_entry_low": pullback_entry_low,
            "pullback_entry_high": pullback_entry_high,
            "pullback_stop_loss": pullback_stop_loss,
            "pullback_target_1": pullback_target_1,
            "pullback_target_2": pullback_target_2,
            "pullback_rrr": pullback_rrr,

            "breakout_stop_loss": breakout_stop_loss,
            "breakout_target_1": breakout_target_1,
            "breakout_target_2": breakout_target_2,
            "breakout_rrr": breakout_rrr,
            "smart_money_percent": smart_money_percent,
            "retail_percent": retail_percent,
        }

    def evaluate_minervini_template(
        self,
        df: pd.DataFrame,
    ) -> dict[str, Any]:
        """
        Basic Minervini-style trend template.

        Requirements:
        1. Close > SMA150
        2. Close > SMA200
        3. SMA50 > SMA150
        4. Close >= 52-week low × 1.30

        The 52-week low is based on the latest 252
        trading sessions where available.
        """

        latest = df.iloc[-1]

        close = latest["Close"]
        sma50 = latest["SMA50"]
        sma150 = latest["SMA150"]
        sma200 = latest["SMA200"]

        one_year_window = df.tail(252)

        fifty_two_week_low = (
            one_year_window["Low"].min()
            if not one_year_window.empty
            else np.nan
        )

        close_above_sma150 = (
            pd.notna(sma150)
            and close > sma150
        )

        close_above_sma200 = (
            pd.notna(sma200)
            and close > sma200
        )

        sma50_above_sma150 = (
            pd.notna(sma50)
            and pd.notna(sma150)
            and sma50 > sma150
        )

        above_30_percent_from_low = (
            pd.notna(fifty_two_week_low)
            and close >= (
                fifty_two_week_low * 1.30
            )
        )

        checks = [
            close_above_sma150,
            close_above_sma200,
            sma50_above_sma150,
            above_30_percent_from_low,
        ]

        passed_checks = sum(checks)

        passed = all(checks)

        failed_rules = []

        if not close_above_sma150:
            failed_rules.append(
                "Close is not above SMA150."
            )

        if not close_above_sma200:
            failed_rules.append(
                "Close is not above SMA200."
            )

        if not sma50_above_sma150:
            failed_rules.append(
                "SMA50 is not above SMA150."
            )

        if not above_30_percent_from_low:
            failed_rules.append(
                "Close is less than 30% above "
                "the 52-week low."
            )

        return {
            "passed": passed,
            "passed_checks": passed_checks,
            "total_checks": 4,
            "close": close,
            "sma50": sma50,
            "sma150": sma150,
            "sma200": sma200,
            "fifty_two_week_low": fifty_two_week_low,
            "close_above_sma150": close_above_sma150,
            "close_above_sma200": close_above_sma200,
            "sma50_above_sma150": sma50_above_sma150,
            "above_30_percent_from_low": (
                above_30_percent_from_low
            ),
            "failed_rules": failed_rules,
        }  


# ============================================================
# EXCEL REPORT BUILDER
# ============================================================

class ExcelReportBuilder:
    def __init__(self) -> None:
        self.workbook = Workbook()
        self.workbook.remove(self.workbook.active)

        self.temporary_chart_files: list[Path] = []

        self.border = Side(
            style="thin",
            color="D9E2F3",
        )

    @staticmethod
    def clean_value(value: Any) -> Any:
        if value is None:
            return None

        if isinstance(value, float) and np.isnan(value):
            return None

        return value

    def set_title(
        self,
        worksheet,
        title: str,
        last_column: int,
    ) -> None:
        worksheet.merge_cells(
            start_row=1,
            start_column=1,
            end_row=1,
            end_column=last_column,
        )

        cell = worksheet.cell(1, 1, title)

        cell.fill = PatternFill(
            "solid",
            fgColor=NAVY,
        )

        cell.font = Font(
            bold=True,
            color=WHITE,
            size=14,
        )

        cell.alignment = Alignment(
            horizontal="center",
            vertical="center",
        )

        worksheet.row_dimensions[1].height = 28
        worksheet.sheet_view.showGridLines = False

    def set_section_title(
        self,
        worksheet,
        row: int,
        title: str,
        last_column: int,
    ) -> None:
        worksheet.merge_cells(
            start_row=row,
            start_column=1,
            end_row=row,
            end_column=last_column,
        )

        cell = worksheet.cell(row, 1, title)

        cell.fill = PatternFill(
            "solid",
            fgColor=BLUE,
        )

        cell.font = Font(
            bold=True,
            color=WHITE,
        )

        cell.alignment = Alignment(
            horizontal="left",
            vertical="center",
        )

    def set_header(
        self,
        worksheet,
        row: int,
        headers: list[str],
    ) -> None:
        for column, header in enumerate(headers, start=1):
            cell = worksheet.cell(
                row=row,
                column=column,
                value=header,
            )

            cell.fill = PatternFill(
                "solid",
                fgColor=NAVY,
            )

            cell.font = Font(
                bold=True,
                color=WHITE,
            )

            cell.alignment = Alignment(
                horizontal="center",
                vertical="center",
                wrap_text=True,
            )

            cell.border = Border(
                bottom=self.border,
            )

    def zebra_rows(
        self,
        worksheet,
        first_row: int,
        last_row: int,
        last_column: int,
    ) -> None:
        for row in range(first_row, last_row + 1):
            if row % 2 == 0:
                for column in range(1, last_column + 1):
                    worksheet.cell(
                        row,
                        column,
                    ).fill = PatternFill(
                        "solid",
                        fgColor=ZEBRA_BLUE,
                    )

    def write_rows(
        self,
        worksheet,
        start_row: int,
        rows: list[list[Any]],
    ) -> int:
        current_row = start_row

        for row_values in rows:
            for column, value in enumerate(
                row_values,
                start=1,
            ):
                worksheet.cell(
                    current_row,
                    column,
                    self.clean_value(value),
                )

            current_row += 1

        return current_row

    def add_dashboard_candlestick_chart(
        self,
        ws,
        ticker: str,
        report_df: pd.DataFrame,
        metrics: dict[str, Any],
    ) -> None:
        if report_df.empty:
            logger.warning(
                "Dashboard chart skipped: no reporting data."
            )
            return

        required_columns = [
            "Date",
            "Open",
            "High",
            "Low",
            "Close",
            "Volume",
        ]

        missing_columns = [
            column
            for column in required_columns
            if column not in report_df.columns
        ]

        if missing_columns:
            logger.warning(
                "Dashboard chart skipped; missing columns: %s",
                missing_columns,
            )
            return

        chart_df = report_df.copy()

        chart_df["Date"] = pd.to_datetime(
            chart_df["Date"]
        )

        chart_df = (
            chart_df
            .sort_values("Date")
            .set_index("Date")
        )

        chart_df = chart_df[
            [
                "Open",
                "High",
                "Low",
                "Close",
                "Volume",
            ]
        ].copy()

        if len(chart_df) < 20:
            logger.warning(
                "Dashboard chart skipped: insufficient records."
            )
            return

        chart_directory = (
            CACHE_DIR
            / "dashboard_charts"
        )

        chart_directory.mkdir(
            parents=True,
            exist_ok=True,
        )

        timestamp = datetime.now().strftime(
            "%Y%m%d_%H%M%S"
        )

        ticker_code = ticker.replace(
            ".JK",
            "",
        )

        chart_path = (
            chart_directory
            / f"{ticker_code}_dashboard_{timestamp}.png"
        )

        add_plots = []

        if "SMA20" in report_df.columns:
            add_plots.append(
                mpf.make_addplot(
                    report_df.set_index("Date")["SMA20"],
                    color="#1F77B4",
                    width=1.0,
                )
            )

        if "SMA50" in report_df.columns:
            add_plots.append(
                mpf.make_addplot(
                    report_df.set_index("Date")["SMA50"],
                    color="#FF7F0E",
                    width=1.0,
                )
            )

        level_definitions = [
            (
                "Pullback Stop",
                metrics.get("pullback_stop_loss"),
                "#C00000",
            ),
            (
                "Pullback Target 1",
                metrics.get("pullback_target_1"),
                "#70AD47",
            ),
            (
                "Breakout Trigger",
                metrics.get("breakout_entry"),
                "#7030A0",
            ),
            (
                "Breakout Target 1",
                metrics.get("breakout_target_1"),
                "#00B0F0",
            ),
            (
                "Breakout Target 2",
                metrics.get("breakout_target_2"),
                "#548235",
            ),
        ]

        horizontal_lines = []
        horizontal_colors = []

        for _, value, color in level_definitions:
            try:
                numeric_value = float(value)

                if pd.notna(numeric_value):
                    horizontal_lines.append(
                        numeric_value
                    )

                    horizontal_colors.append(
                        color
                    )

            except (
                TypeError,
                ValueError,
            ):
                continue

        chart_style = mpf.make_mpf_style(
            base_mpf_style="yahoo",
            marketcolors=mpf.make_marketcolors(
                up="#26A69A",
                down="#EF5350",
                edge="inherit",
                wick="inherit",
                volume="inherit",
            ),
            facecolor="white",
            gridcolor="#D9E1E8",
            gridstyle="-",
            rc={
                "font.size": 8,
            },
        )

        plot_arguments = {
            "type": "candle",
            "style": chart_style,
            "volume": True,
            "addplot": add_plots,
            "figratio": (18, 10),
            "figscale": 0.78,
            "title": (
                f"{ticker} — 6-Month Price Structure"
            ),
            "ylabel": "Price",
            "ylabel_lower": "Volume",
            "tight_layout": True,
            "savefig": {
                "fname": str(chart_path),
                "dpi": 150,
                "bbox_inches": "tight",
            },
        }

        if horizontal_lines:
            plot_arguments["hlines"] = {
                "hlines": horizontal_lines,
                "colors": horizontal_colors,
                "linewidths": 0.9,
                "linestyle": "--",
            }

        try:
            mpf.plot(
                chart_df,
                **plot_arguments,
            )

            chart_image = OpenpyxlImage(
                str(chart_path)
            )

            chart_image.width = 850
            chart_image.height = 470

            ws.add_image(
                chart_image,
                "E11",
            )

            self.temporary_chart_files.append(
                chart_path
            )

            logger.info(
                "Embedded dashboard chart: %s",
                chart_path.name,
            )

        except Exception as error:
            logger.warning(
                "Dashboard chart generation failed: %s",
                error,
            )

    def build_dashboard(
        self,
        ticker: str,
        fundamentals: dict[str, Any],
        metrics: dict[str, Any],
        generated_at: str,
        broker_data_source: str,
        api_calls_used: int,
        broker_lookback_days: int | None,
        report_df: pd.DataFrame,
    ) -> None:
        ws = self.workbook.create_sheet(
            "1. Decision Dashboard"
        )

        self.set_title(
            ws,
            "Automated IDX Swing Trading Decision Dashboard",
            8,
        )

        metadata = [
            ["Ticker", ticker],
            ["Company", fundamentals["name"]],
            ["Generated At", generated_at],
            ["Data Horizon", "6 Months Daily Data"],
        ]

        for row, values in enumerate(metadata, start=3):
            ws.cell(row, 1, values[0]).font = Font(bold=True)
            ws.cell(row, 2, values[1])

        # ----------------------------------------------------
        # Visual dashboard cards
        # ----------------------------------------------------

        def create_dashboard_card(
            title: str,
            value: str,
            start_column: int,
            start_row: int,
            fill_color: str,
            value_color: str = WHITE,
        ) -> None:
            end_column = start_column + 1

            ws.merge_cells(
                start_row=start_row,
                start_column=start_column,
                end_row=start_row,
                end_column=end_column,
            )

            ws.merge_cells(
                start_row=start_row + 1,
                start_column=start_column,
                end_row=start_row + 2,
                end_column=end_column,
            )

            title_cell = ws.cell(
                start_row,
                start_column,
                title,
            )

            title_cell.fill = PatternFill(
                "solid",
                fgColor=NAVY,
            )

            title_cell.font = Font(
                bold=True,
                color=WHITE,
                size=10,
            )

            title_cell.alignment = Alignment(
                horizontal="center",
                vertical="center",
            )

            value_cell = ws.cell(
                start_row + 1,
                start_column,
                value,
            )

            value_cell.fill = PatternFill(
                "solid",
                fgColor=fill_color,
            )

            value_cell.font = Font(
                bold=True,
                color=value_color,
                size=12,
            )

            value_cell.alignment = Alignment(
                horizontal="center",
                vertical="center",
                wrap_text=True,
            )

        # ----------------------------------------------------
        # Master Signal card color
        # ----------------------------------------------------

        if metrics["decision"] in [
            "STRONG BUY / INVEST",
            "BUY CANDIDATE - FLOW UNVERIFIED",
        ]:
            master_signal_fill = GREEN
            master_signal_text = GREEN_TEXT

        elif metrics["decision"] in [
            "HOLD / WATCHLIST",
            "HOLD / WAIT FOR PULLBACK",
        ]:
            master_signal_fill = YELLOW
            master_signal_text = YELLOW_TEXT

        else:
            master_signal_fill = RED
            master_signal_text = RED_TEXT

        # ----------------------------------------------------
        # Trend card color
        # ----------------------------------------------------

        if metrics["minervini_passed"]:
            trend_fill = GREEN
            trend_text = GREEN_TEXT
        else:
            trend_fill = RED
            trend_text = RED_TEXT

        # ----------------------------------------------------
        # Flow card color
        # ----------------------------------------------------

        if (
            metrics["broker_data_available"]
            and metrics["foreign_flow_data_available"]
        ):
            flow_fill = GREEN
            flow_text = GREEN_TEXT
        else:
            flow_fill = YELLOW
            flow_text = YELLOW_TEXT

        # ----------------------------------------------------
        # Trade Execution card color
        # ----------------------------------------------------

        if (
            metrics["extension_risk_status"]
            == "BUYING CLIMAX / SEVERELY OVEREXTENDED"
        ):
            trade_fill = RED
            trade_text = RED_TEXT

        elif (
            metrics["extension_risk_status"]
            == "EXTENDED"
        ):
            trade_fill = YELLOW
            trade_text = YELLOW_TEXT

        else:
            trade_fill = GREEN
            trade_text = GREEN_TEXT

        # ----------------------------------------------------
        # Create four cards
        # ----------------------------------------------------

        create_dashboard_card(
            "MASTER SIGNAL",
            (
                f"{metrics['decision']}\n"
                f"{metrics['normalized_score']:.1f}/100"
            ),
            5,
            3,
            master_signal_fill,
            master_signal_text,
        )

        create_dashboard_card(
            "TREND STRUCTURE",
            (
                f"{metrics['wyckoff_phase']}\n"
                f"Minervini: "
                f"{metrics['minervini_passed_checks']}"
                f"/{metrics['minervini_total_checks']}"
            ),
            7,
            3,
            trend_fill,
            trend_text,
        )

        create_dashboard_card(
            "FLOW CONFIRMATION",
            (
                "CONNECTED"
                if (
                    metrics["broker_data_available"]
                    and metrics["foreign_flow_data_available"]
                )
                else "UNAVAILABLE"
            ),
            5,
            7,
            flow_fill,
            flow_text,
        )

        create_dashboard_card(
            "TRADE EXECUTION",
            (
                f"Extension: "
                f"{metrics['extension_risk_status']}\n"
                f"RRR: "
                f"{metrics['pullback_rrr']:.2f}"
                if pd.notna(metrics["pullback_rrr"])
                else (
                    "Extension: "
                    f"{metrics['extension_risk_status']}\n"
                    "RRR: N/A"
                )
            ),
            7,
            7,
            trade_fill,
            trade_text,
        )

        self.set_section_title(
            ws,
            8,
            "Master Decision",
            4,
        )

        broker_data_status = (
            "CONNECTED"
            if metrics["broker_data_available"]
            else "NOT CONNECTED"
        )

        foreign_flow_status = (
            "CONNECTED"
            if metrics["foreign_flow_data_available"]
            else "NOT CONNECTED"
        )

        decision_rows = [
            [
                "Decision",
                metrics["decision"],
                "Confidence",
                metrics["confidence"],
            ],
            [
                "Normalized Score",
                f"{metrics['normalized_score']:.1f} / 100",
                "Data Coverage",
                f"{metrics['data_coverage_percent']:.0f}%",
            ],
            [
                "Raw Score",
                (
                    f"{metrics['raw_score']} / "
                    f"{metrics['eligible_max_score']} "
                    "eligible points"
                ),
                "Wyckoff Phase",
                metrics["wyckoff_phase"],
            ],
            [
                "Broker Summary Data",
                broker_data_status,
                "Foreign Flow Data",
                foreign_flow_status,
            ],
            [
                "Gorengan Risk",
                metrics["risk_label"],
                "Risk Points",
                f"{metrics['risk_points']} / 4",
            ],
            [
                "Trend Bias",
                (
                    f"SMA20: {metrics['trend_sma20']} | "
                    f"SMA50: {metrics['trend_sma50']}"
                ),
                "IHSG Relative Strength",
                metrics["relative_strength_signal"],
            ],
            [
                "Minervini Trend Template",
                (
                    "PASSED"
                    if metrics["minervini_passed"]
                    else "FAILED"
                ),
                "Structural Checks",
                (
                    f"{metrics['minervini_passed_checks']}"
                    f"/{metrics['minervini_total_checks']}"
                ),
            ],
            [
                "Extension Risk",
                metrics["extension_risk_status"],
                "Entry Timing",
                metrics["extension_risk_reason"],
            ],
        ]

        self.write_rows(
            ws,
            9,
            decision_rows,
        )

        decision_cell = ws["B9"]

        if metrics["decision"] in [
            "STRONG BUY / INVEST",
            "BUY CANDIDATE - FLOW UNVERIFIED",
        ]:
            decision_cell.fill = PatternFill(
                "solid",
                fgColor=GREEN,
            )
            decision_cell.font = Font(
                bold=True,
                color=GREEN_TEXT,
            )

        elif metrics["decision"] == "HOLD / WATCHLIST":
            decision_cell.fill = PatternFill(
                "solid",
                fgColor=YELLOW,
            )
            decision_cell.font = Font(
                bold=True,
                color=YELLOW_TEXT,
            )

        else:
            decision_cell.fill = PatternFill(
                "solid",
                fgColor=RED,
            )
            decision_cell.font = Font(
                bold=True,
                color=RED_TEXT,
            )

        risk_cell = ws["B13"]

        # ----------------------------------------------------
        # Minervini Trend Template status badge
        # ----------------------------------------------------

        minervini_cell = ws["B15"]

        if metrics["minervini_passed"]:
            minervini_cell.fill = PatternFill(
                "solid",
                fgColor=GREEN,
            )

            minervini_cell.font = Font(
                bold=True,
                color=GREEN_TEXT,
            )

        else:
            minervini_cell.fill = PatternFill(
                "solid",
                fgColor=RED,
            )

            minervini_cell.font = Font(
                bold=True,
                color=RED_TEXT,
            )

        # ----------------------------------------------------
        # Extension Risk status badge
        # ----------------------------------------------------

        extension_cell = ws["B16"]

        if (
            metrics["extension_risk_status"]
            == "NORMAL"
        ):
            extension_cell.fill = PatternFill(
                "solid",
                fgColor=GREEN,
            )

            extension_cell.font = Font(
                bold=True,
                color=GREEN_TEXT,
            )

        elif (
            metrics["extension_risk_status"]
            == "EXTENDED"
        ):
            extension_cell.fill = PatternFill(
                "solid",
                fgColor=YELLOW,
            )

            extension_cell.font = Font(
                bold=True,
                color=YELLOW_TEXT,
            )

        else:
            extension_cell.fill = PatternFill(
                "solid",
                fgColor=RED,
            )

            extension_cell.font = Font(
                bold=True,
                color=RED_TEXT,
            )

        if metrics["risk_points"] == 0:
            risk_cell.fill = PatternFill(
                "solid",
                fgColor=GREEN,
            )
        elif metrics["risk_points"] <= 2:
            risk_cell.fill = PatternFill(
                "solid",
                fgColor=YELLOW,
            )
        else:
            risk_cell.fill = PatternFill(
                "solid",
                fgColor=RED,
            )

        self.set_section_title(
            ws,
            17,
            "Why This Decision?",
            4,
        )

        self.set_header(
            ws,
            18,
            [
                "Pillar",
                "Result",
                "Points",
                "Evidence",
            ],
        )

        evidence_start = 19

        for row, evidence in enumerate(
            metrics["evidence"],
            start=evidence_start,
        ):
            ws.cell(row, 1, evidence[0])
            ws.cell(row, 2, evidence[1])
            ws.cell(row, 3, evidence[2])
            ws.cell(row, 4, evidence[3])

            result_cell = ws.cell(row, 2)

            if evidence[1] == "Positive":
                result_cell.fill = PatternFill(
                    "solid",
                    fgColor=GREEN,
                )

            elif (
                "Negative" in evidence[1]
                or "Downgraded" in evidence[1]
            ):
                result_cell.fill = PatternFill(
                    "solid",
                    fgColor=RED,
                )

            else:
                result_cell.fill = PatternFill(
                    "solid",
                    fgColor=YELLOW,
                )

        risk_start = evidence_start + len(
            metrics["evidence"]
        ) + 2

        self.set_section_title(
            ws,
            risk_start,
            "Risk Flags",
            4,
        )

        for row, reason in enumerate(
            metrics["risk_reasons"],
            start=risk_start + 1,
        ):
            ws.cell(row, 1, "Risk Check")
            ws.cell(row, 2, reason)

        # ----------------------------------------------------
        # Data Quality & Limitations
        # ----------------------------------------------------

        data_quality_start = (
            risk_start
            + len(metrics["risk_reasons"])
            + 3
        )

        self.set_section_title(
            ws,
            data_quality_start,
            "Data Quality & Limitations",
            4,
        )

        operating_cash_flow_status = (
            "CONNECTED"
            if fundamentals["operating_cash_flow"] is not None
            else "UNAVAILABLE"
        )

        broker_status = (
            "CONNECTED"
            if metrics["broker_data_available"]
            else "NOT CONNECTED"
        )

        foreign_flow_status = (
            "CONNECTED"
            if metrics["foreign_flow_data_available"]
            else "NOT CONNECTED"
        )

        cache_status = (
            "CACHE USED"
            if (
                metrics["broker_data_available"]
                and api_calls_used == 0
                and broker_data_source != "Not Connected"
            )
            else (
                f"{api_calls_used} API REQUEST(S)"
                if metrics["broker_data_available"]
                else "NO API REQUEST"
            )
        )

        lookback_status = (
            f"{broker_lookback_days}-DAY AGGREGATED WINDOW"
            if broker_lookback_days is not None
            else "NOT APPLICABLE"
        )

        quality_rows = [
            [
                "Price & Volume",
                "Yahoo Finance",
                "CONNECTED",
                "Daily OHLCV data for the latest six-month horizon.",
            ],
            [
                "Fundamentals",
                "Yahoo Finance",
                "PARTIAL"
                if operating_cash_flow_status == "UNAVAILABLE"
                else "CONNECTED",
                (
                    "Operating Cash Flow is unavailable and "
                    "excluded from the scoring denominator."
                    if operating_cash_flow_status == "UNAVAILABLE"
                    else "Income, balance-sheet, and cash-flow "
                    "fields are available."
                ),
            ],
            [
                "Broker Summary",
                broker_data_source,
                broker_status,
                (
                    f"{lookback_status}. Broker activity is "
                    "a transaction-routing proxy, not verified "
                    "beneficial ownership."
                ),
            ],
            [
                "Foreign Investor Flow",
                broker_data_source,
                foreign_flow_status,
                (
                    "Foreign-only broker buy value minus sell "
                    "value across the selected aggregation window."
                ),
            ],
            [
                "API Usage / Cache",
                "Index Alpha Cache",
                cache_status,
                (
                    "Cached responses reduce API usage and prevent "
                    "repeated requests for the same ticker/date window."
                ),
            ],
            [
                "Decision Score Coverage",
                "Dynamic Scoring",
                f"{metrics['data_coverage_percent']:.0f}%",
                (
                    f"{metrics['eligible_max_score']} of 100 "
                    "maximum points are eligible. Missing data "
                    "is excluded rather than treated as negative."
                ),
            ],
        ]

        self.set_header(
            ws,
            data_quality_start + 1,
            [
                "Data Category",
                "Source",
                "Status",
                "Meaning / Limitation",
            ],
        )

        for row_number, row_values in enumerate(
            quality_rows,
            start=data_quality_start + 2,
        ):
            for column_number, value in enumerate(
                row_values,
                start=1,
            ):
                ws.cell(
                    row_number,
                    column_number,
                    value=value,
                )

            status_cell = ws.cell(
                row_number,
                3,
            )

            status_text = str(
                status_cell.value
            ).upper()

            if (
                "CONNECTED" in status_text
                or "CACHE USED" in status_text
            ):
                status_cell.fill = PatternFill(
                    "solid",
                    fgColor=GREEN,
                )

            elif (
                "UNAVAILABLE" in status_text
                or "NOT CONNECTED" in status_text
                or "PARTIAL" in status_text
            ):
                status_cell.fill = PatternFill(
                    "solid",
                    fgColor=YELLOW,
                )

        ws.column_dimensions["A"].width = 25
        ws.column_dimensions["B"].width = 34
        ws.column_dimensions["C"].width = 18
        ws.column_dimensions["D"].width = 65
        ws.column_dimensions["E"].width = 18
        ws.column_dimensions["F"].width = 18
        ws.column_dimensions["G"].width = 18
        ws.column_dimensions["H"].width = 18

        ws.freeze_panes = "A9"

        self.add_dashboard_candlestick_chart(
            ws,
            ticker,
            report_df,
            metrics,
        )

    def build_trade_plan(
        self,
        ticker: str,
        metrics: dict[str, Any],
    ) -> None:
        ws = self.workbook.create_sheet(
            "2. Trade Plan"
        )

        self.set_title(
            ws,
            f"Rule-Based Swing Trade Plan: {ticker}",
            4,
        )

        # ----------------------------------------------------
        # Pullback plan
        # ----------------------------------------------------

        self.set_section_title(
            ws,
            3,
            "Scenario A: Pullback Swing Plan",
            4,
        )

        self.set_header(
            ws,
            4,
            [
                "Trade Parameter",
                "Price / Value",
                "Meaning",
                "Action Rule",
            ],
        )

        pullback_rrr_text = (
            f"{metrics['pullback_rrr']:.2f}"
            if pd.notna(metrics["pullback_rrr"])
            else "N/A"
        )

        pullback_rows = [
            [
                "Current Close",
                metrics["latest_close"],
                "Latest available closing price.",
                "Use only as a reference price.",
            ],
            [
                "Preferred Entry Zone",
                (
                    f"{metrics['pullback_entry_low']:,.2f} "
                    f"to {metrics['pullback_entry_high']:,.2f}"
                    if pd.notna(
                        metrics["pullback_entry_low"]
                    )
                    else "N/A"
                ),
                "ATR-based pullback zone below or near current price.",
                "Avoid chasing materially above this zone.",
            ],
            [
                "Pullback Stop Loss",
                metrics["pullback_stop_loss"],
                "Approximately 2 × ATR14 below current close.",
                "Exit if price closes below the planned risk level.",
            ],
            [
                "Pullback Target 1",
                metrics["pullback_target_1"],
                "First resistance or at least 1.5 × ATR upside.",
                "Consider partial profit-taking.",
            ],
            [
                "Pullback Target 2",
                metrics["pullback_target_2"],
                "Six-month high or at least 3 × ATR upside.",
                "Use a trailing stop if momentum remains strong.",
            ],
            [
                "Pullback RRR",
                pullback_rrr_text,
                "Potential reward divided by pre-defined risk.",
                "Prefer RRR above 2.00 before entering.",
            ],
        ]

        self.write_rows(
            ws,
            5,
            pullback_rows,
        )

        for row in [5, 7, 8, 9]:
            ws.cell(
                row,
                2,
            ).number_format = "#,##0.00"

        # ----------------------------------------------------
        # Breakout plan
        # ----------------------------------------------------

        self.set_section_title(
            ws,
            13,
            "Scenario B: Confirmed Breakout Plan",
            4,
        )

        self.set_header(
            ws,
            14,
            [
                "Trade Parameter",
                "Price / Value",
                "Meaning",
                "Action Rule",
            ],
        )

        breakout_rrr_text = (
            f"{metrics['breakout_rrr']:.2f}"
            if pd.notna(metrics["breakout_rrr"])
            else "N/A"
        )

        breakout_rows = [
            [
                "Breakout Trigger",
                metrics["breakout_entry"],
                "20-day resistance level.",
                "Enter only after a daily close above this level.",
            ],
            [
                "Volume Confirmation",
                "Volume Ratio > 1.20x",
                "Higher volume improves breakout quality.",
                "Avoid weak breakouts with low volume.",
            ],
            [
                "Breakout Stop Loss",
                metrics["breakout_stop_loss"],
                "Approximately 1.5 × ATR14 below breakout entry.",
                "Exit if breakout fails and price closes below stop.",
            ],
            [
                "Breakout Target 1",
                metrics["breakout_target_1"],
                "Breakout entry plus 1.5 × ATR14.",
                "Consider partial profit-taking.",
            ],
            [
                "Breakout Target 2",
                metrics["breakout_target_2"],
                "Breakout entry plus 3.0 × ATR14.",
                "Hold only if trend, volume, and momentum remain supportive.",
            ],
            [
                "Breakout RRR",
                breakout_rrr_text,
                "Potential reward divided by planned breakout risk.",
                "Prefer RRR above 2.00 before entering.",
            ],
        ]

        self.write_rows(
            ws,
            15,
            breakout_rows,
        )

        for row in [15, 17, 18, 19]:
            ws.cell(
                row,
                2,
            ).number_format = "#,##0.00"

        # ----------------------------------------------------
        # Setup quality assessment
        # ----------------------------------------------------

        self.set_section_title(
            ws,
            23,
            "Setup Quality Assessment",
            4,
        )

        pullback_valid = (
            pd.notna(metrics["pullback_rrr"])
            and metrics["pullback_rrr"] >= 2.0
        )

        breakout_valid = (
            pd.notna(metrics["breakout_rrr"])
            and metrics["breakout_rrr"] >= 2.0
        )

        pullback_status = (
            "VALID SETUP"
            if pullback_valid
            else "NO VALID SETUP"
        )

        breakout_status = (
            "VALID SETUP"
            if breakout_valid
            else "NO VALID SETUP"
        )

        setup_rows = [
            [
                "Pullback Setup Status",
                pullback_status,
                (
                    f"RRR: {metrics['pullback_rrr']:.2f}"
                    if pd.notna(metrics["pullback_rrr"])
                    else "ATR data unavailable."
                ),
                (
                    "Potentially acceptable only if the trend thesis remains valid."
                    if pullback_valid
                    else "Upside relative to planned risk is insufficient."
                ),
            ],
            [
                "Breakout Setup Status",
                breakout_status,
                (
                    f"RRR: {metrics['breakout_rrr']:.2f}"
                    if pd.notna(metrics["breakout_rrr"])
                    else "ATR data unavailable."
                ),
                (
                    "Wait for a daily close above resistance and volume confirmation."
                    if breakout_valid
                    else "Do not enter unless the trade plan improves."
                ),
            ],
        ]

        self.write_rows(
            ws,
            24,
            setup_rows,
        )

        for row in [24, 25]:
            status_cell = ws.cell(row, 2)

            if status_cell.value == "VALID SETUP":
                status_cell.fill = PatternFill(
                    "solid",
                    fgColor=GREEN,
                )
                status_cell.font = Font(
                    bold=True,
                    color=GREEN_TEXT,
                )
            else:
                status_cell.fill = PatternFill(
                    "solid",
                    fgColor=RED,
                )
                status_cell.font = Font(
                    bold=True,
                    color=RED_TEXT,
                )

        # ----------------------------------------------------
        # Risk disclaimer
        # ----------------------------------------------------

        self.set_section_title(
            ws,
            28,
            "Important Note",
            4,
        )

        ws.merge_cells(
            start_row=29,
            start_column=1,
            end_row=32,
            end_column=4,
        )

        note_cell = ws.cell(
            29,
            1,
            (
                "This workbook provides a rule-based analytical "
                "trade plan, not investment advice or a guaranteed "
                "recommendation. Prices can gap through stop-loss levels. "
                "Validate liquidity, corporate actions, market conditions, "
                "and your personal maximum risk before entering a trade."
            ),
        )

        note_cell.fill = PatternFill(
            "solid",
            fgColor=ORANGE,
        )

        note_cell.alignment = Alignment(
            wrap_text=True,
            vertical="top",
        )

        ws.column_dimensions["A"].width = 27
        ws.column_dimensions["B"].width = 27
        ws.column_dimensions["C"].width = 44
        ws.column_dimensions["D"].width = 54

        ws.freeze_panes = "A5"

    def build_technicals(
        self,
        ticker: str,
        metrics: dict[str, Any],
    ) -> None:
        ws = self.workbook.create_sheet(
            "3. Price & Technicals"
        )

        self.set_title(
            ws,
            f"Price, Trend, Momentum & Volatility: {ticker}",
            4,
        )

        self.set_header(
            ws,
            3,
            [
                "Indicator",
                "Current Value",
                "Signal",
                "Interpretation",
            ],
        )

        rows = [
            [
                "Close Price",
                metrics["latest_close"],
                "Reference",
                "Latest available close.",
            ],
            [
                "SMA20",
                metrics["sma20"],
                metrics["trend_sma20"],
                "Short-term trend reference.",
            ],
            [
                "SMA50",
                metrics["sma50"],
                metrics["trend_sma50"],
                "Medium-term trend reference.",
            ],
            [
                "EMA9 vs EMA21",
                (
                    f"{metrics['ema9']:,.2f} / "
                    f"{metrics['ema21']:,.2f}"
                ),
                metrics["ema_signal"],
                "Fast moving-average momentum relationship.",
            ],
            [
                "RSI14",
                metrics["rsi14"],
                (
                    "CONSTRUCTIVE"
                    if (
                        pd.notna(metrics["rsi14"])
                        and 45 <= metrics["rsi14"] <= 65
                    )
                    else "EXTENDED / WEAK"
                ),
                "45 to 65 is a preferred constructive swing-momentum zone.",
            ],
            [
                "MACD",
                metrics["macd"],
                metrics["macd_signal"],
                "MACD above signal line indicates improving momentum.",
            ],
            [
                "ADX14",
                metrics["adx14"],
                metrics["adx_signal"],
                "Trend-strength gauge. Above 25 suggests a stronger trend.",
            ],
            [
                "Volume Ratio",
                metrics["volume_ratio"],
                (
                    "HIGH PARTICIPATION"
                    if (
                        pd.notna(metrics["volume_ratio"])
                        and metrics["volume_ratio"] > 1.5
                    )
                    else "NORMAL / LOW"
                ),
                "Compares current volume with 20-day average volume.",
            ],
            [
                "ATR14",
                metrics["atr14"],
                "VOLATILITY",
                "Average daily movement used for stop-loss and target planning.",
            ],
            [
                "ATR %",
                metrics["atr_percent"],
                "VOLATILITY",
                "ATR14 expressed as a percentage of current price.",
            ],
            [
                "Bollinger Upper",
                metrics["bb_upper"],
                "VOLATILITY BAND",
                "Upper two-standard-deviation price band.",
            ],
            [
                "Bollinger Middle",
                metrics["bb_middle"],
                "20D MEAN",
                "20-day moving-average center line.",
            ],
            [
                "Bollinger Lower",
                metrics["bb_lower"],
                "VOLATILITY BAND",
                "Lower two-standard-deviation price band.",
            ],
            [
                "20D Support",
                metrics["support"],
                "SUPPORT",
                "Lowest low across the latest 20 trading days.",
            ],
            [
                "20D Resistance",
                metrics["resistance"],
                "RESISTANCE",
                "Highest high across the latest 20 trading days.",
            ],
            [
                "Relative Strength vs IHSG",
                metrics["relative_strength_signal"],
                "BENCHMARK",
                "Shows whether the stock is outperforming or underperforming IHSG over 20 days.",
            ],
        ]

        self.write_rows(ws, 4, rows)

        for row in range(4, ws.max_row + 1):
            ws.cell(row, 2).number_format = "#,##0.00"

        ws["B13"].number_format = "0.00%"
        ws["B11"].number_format = '0.00"x"'

        ws.column_dimensions["A"].width = 28
        ws.column_dimensions["B"].width = 25
        ws.column_dimensions["C"].width = 28
        ws.column_dimensions["D"].width = 58

        daily_ws = self.workbook[
            "8. Raw Daily Data"
        ]

        chart = LineChart()
        chart.title = "Daily Closing Price"
        chart.y_axis.title = "Price (IDR)"
        chart.x_axis.title = "Date"
        chart.height = 8
        chart.width = 18

        dates = Reference(
            daily_ws,
            min_col=1,
            min_row=4,
            max_row=daily_ws.max_row,
        )

        close_price_data = Reference(
            daily_ws,
            min_col=5,
            min_row=3,
            max_row=daily_ws.max_row,
        )

        sma20_data = Reference(
            daily_ws,
            min_col=15,
            min_row=3,
            max_row=daily_ws.max_row,
        )

        sma50_data = Reference(
            daily_ws,
            min_col=16,
            min_row=3,
            max_row=daily_ws.max_row,
        )

        pullback_stop_data = Reference(
            daily_ws,
            min_col=35,
            min_row=3,
            max_row=daily_ws.max_row,
        )

        pullback_target_data = Reference(
            daily_ws,
            min_col=36,
            min_row=3,
            max_row=daily_ws.max_row,
        )

        breakout_trigger_data = Reference(
            daily_ws,
            min_col=37,
            min_row=3,
            max_row=daily_ws.max_row,
        )

        breakout_target_1_data = Reference(
            daily_ws,
            min_col=38,
            min_row=3,
            max_row=daily_ws.max_row,
        )

        breakout_target_2_data = Reference(
            daily_ws,
            min_col=39,
            min_row=3,
            max_row=daily_ws.max_row,
        )

        chart.add_data(
            close_price_data,
            titles_from_data=True,
        )

        chart.add_data(
            sma20_data,
            titles_from_data=True,
        )

        chart.add_data(
            sma50_data,
            titles_from_data=True,
        )

        chart.add_data(
            pullback_stop_data,
            titles_from_data=True,
        )

        chart.add_data(
            pullback_target_data,
            titles_from_data=True,
        )

        chart.add_data(
            breakout_trigger_data,
            titles_from_data=True,
        )

        chart.add_data(
            breakout_target_1_data,
            titles_from_data=True,
        )

        chart.add_data(
            breakout_target_2_data,
            titles_from_data=True,
        )

        chart.set_categories(dates)

        chart.legend.position = "r"

        chart.series[0].graphicalProperties.line.solidFill = "1F4E78"
        chart.series[0].graphicalProperties.line.width = 30000

        chart.series[1].graphicalProperties.line.solidFill = "70AD47"
        chart.series[2].graphicalProperties.line.solidFill = "ED7D31"

        chart.series[3].graphicalProperties.line.solidFill = "C00000"
        chart.series[4].graphicalProperties.line.solidFill = "70AD47"

        chart.series[5].graphicalProperties.line.solidFill = "7030A0"
        chart.series[6].graphicalProperties.line.solidFill = "5B9BD5"
        chart.series[7].graphicalProperties.line.solidFill = "4472C4"

        ws.add_chart(chart, "A23")
        ws.freeze_panes = "A4"

    def build_fundamentals(
        self,
        fundamentals: dict[str, Any],
    ) -> None:
        ws = self.workbook.create_sheet(
            "4. Fundamentals"
        )

        self.set_title(
            ws,
            "Company Profile, Valuation & Balance Sheet",
            4,
        )

        self.set_header(
            ws,
            3,
            [
                "Section",
                "Metric",
                "Value",
                "Meaning",
            ],
        )

        rows = [
            ["Profile", "Ticker", fundamentals["ticker"], "IDX ticker symbol."],
            ["Profile", "Company Name", fundamentals["name"], "Issuer name from Yahoo Finance."],
            ["Profile", "Sector", fundamentals["sector"], "Business sector classification."],
            ["Profile", "Industry", fundamentals["industry"], "Business industry classification."],
            ["Valuation", "Market Capitalization", fundamentals["market_cap"], "Approximate equity market value."],
            ["Valuation", "Enterprise Value", fundamentals["enterprise_value"], "Equity value plus debt less cash."],
            ["Valuation", "Trailing P/E", fundamentals["trailing_pe"], "Price relative to trailing earnings."],
            ["Valuation", "Forward P/E", fundamentals["forward_pe"], "Price relative to expected earnings."],
            ["Valuation", "Price-to-Book", fundamentals["price_to_book"], "Price relative to book value."],
            ["Balance Sheet", "Cash & Equivalents", fundamentals["cash"], "Short-term liquidity buffer."],
            ["Debt", "Short-Term Debt", fundamentals["short_term_debt"], "Obligations due within one year."],
            ["Debt", "Long-Term Debt", fundamentals["long_term_debt"], "Longer-term financing obligations."],
            ["Debt", "Total Debt", fundamentals["total_debt"], "Short-term plus long-term debt."],
            ["Debt", "Net Debt", fundamentals["net_debt"], "Total debt less available cash."],
            ["Leverage", "Debt-to-Equity", fundamentals["debt_to_equity"], "Debt level relative to shareholder equity."],
            ["Liquidity", "Current Ratio", fundamentals["current_ratio"], "Current assets divided by current liabilities."],
            ["Liquidity", "Quick Ratio", fundamentals["quick_ratio"], "More conservative short-term liquidity ratio."],
            ["Profitability", "Net Income", fundamentals["net_income"], "Latest reported net profit or loss."],
            ["Cash Flow", "Operating Cash Flow", fundamentals["operating_cash_flow"], "Cash generated from core operations."],
        ]

        self.write_rows(ws, 4, rows)
        self.zebra_rows(ws, 4, ws.max_row, 4)

        for row in range(4, ws.max_row + 1):
            ws.cell(row, 3).number_format = "#,##0.00"

        ws.column_dimensions["A"].width = 18
        ws.column_dimensions["B"].width = 30
        ws.column_dimensions["C"].width = 26
        ws.column_dimensions["D"].width = 50

        ws.freeze_panes = "A4"

    def build_bandarmology(
        self,
        df: pd.DataFrame,
        broker_data_source: str,
        lookback_days: int | None,
    ) -> None:
        ws = self.workbook.create_sheet(
            "5. Bandarmology & Flow"
        )

        self.set_title(
            ws,
            "Bandarmology & Foreign Investor Flow",
            8,
        )

        latest_date = pd.to_datetime(
            df["Date"].max()
        ).normalize()

        has_broker_data = (
            df["Net Foreign Flow"]
            .notna()
            .any()
        )

        if (
            lookback_days is not None
            and lookback_days > 0
        ):
            window_start = latest_date - pd.Timedelta(
                days=lookback_days - 1
            )

            window_label = (
                f"{window_start:%Y-%m-%d} "
                f"to {latest_date:%Y-%m-%d}"
            )

            aggregation_label = (
                f"{lookback_days}-Calendar-Day "
                "Aggregated Window"
            )
        else:
            window_label = "N/A"
            aggregation_label = "No Broker Window Selected"

        self.set_section_title(
            ws,
            3,
            "Data Coverage & Interpretation",
            8,
        )

        metadata_rows = [
            [
                "Data Source",
                broker_data_source,
                "Market",
                "Regular Market (RG)",
            ],
            [
                "Broker Summary Window",
                window_label,
                "Aggregation",
                aggregation_label,
            ],
            [
                "Broker Summary Status",
                (
                    "CONNECTED"
                    if has_broker_data
                    else "NOT CONNECTED"
                ),
                "Foreign Investor Flow",
                (
                    "CONNECTED"
                    if has_broker_data
                    else "NOT CONNECTED"
                ),
            ],
            [
                "Interpretation",
                (
                    "Broker-flow proxy. It represents "
                    "aggregated broker transactions, not "
                    "verified beneficial ownership."
                ),
                "Scoring Use",
                (
                    "Positive Top-3 broker net buying and "
                    "positive foreign investor net flow can "
                    "each contribute to the decision score."
                ),
            ],
        ]

        self.write_rows(
            ws,
            4,
            metadata_rows,
        )

        for row in range(4, 8):
            ws.cell(row, 1).font = Font(bold=True)
            ws.cell(row, 3).font = Font(bold=True)

        status_fill = (
            GREEN
            if has_broker_data
            else YELLOW
        )

        for cell_address in ["B6", "D6"]:
            ws[cell_address].fill = PatternFill(
                "solid",
                fgColor=status_fill,
            )

            ws[cell_address].font = Font(
                bold=True,
            )

        self.set_section_title(
            ws,
            10,
            "Latest Aggregated Broker Summary",
            8,
        )

        self.set_header(
            ws,
            11,
            [
                "Window End Date",
                "Foreign Net Flow Proxy (IDR)",
                "Top 3 Accumulating Brokers",
                "Top 3 Distributing Brokers",
                "Top 3 Net Buy Volume",
                "Top 1 Broker Net Buy Volume",
                "Flow Interpretation",
                "Data Note",
            ],
        )

        latest = df.iloc[-1]

        foreign_flow = latest["Net Foreign Flow"]
        top3_net_buy_volume = latest[
            "Top 3 Net Buy Volume"
        ]

        top1_buy_volume = latest[
            "Top 1 Buy Volume"
        ]

        if pd.isna(foreign_flow):
            flow_interpretation = (
                "N/A - Foreign data unavailable"
            )

        elif foreign_flow > 0:
            flow_interpretation = (
                "POSITIVE FOREIGN ACCUMULATION"
            )

        elif foreign_flow < 0:
            flow_interpretation = (
                "FOREIGN DISTRIBUTION"
            )

        else:
            flow_interpretation = "NEUTRAL"

        data_note = (
            "Values are aggregated across the selected "
            "broker-summary window."
            if has_broker_data
            else (
                "No local broker data or Index Alpha API "
                "response was available."
            )
        )

        summary_values = [
            latest["Date"],
            foreign_flow,
            latest["Top 3 Accumulating Brokers"],
            latest["Top 3 Distributing Brokers"],
            top3_net_buy_volume,
            top1_buy_volume,
            flow_interpretation,
            data_note,
        ]

        for column, value in enumerate(
            summary_values,
            start=1,
        ):
            ws.cell(
                12,
                column,
                self.clean_value(value),
            )

        ws.cell(12, 1).number_format = "yyyy-mm-dd"
        ws.cell(12, 2).number_format = "#,##0"
        ws.cell(12, 5).number_format = "#,##0"
        ws.cell(12, 6).number_format = "#,##0"

        if (
            pd.notna(foreign_flow)
            and foreign_flow > 0
        ):
            ws["G12"].fill = PatternFill(
                "solid",
                fgColor=GREEN,
            )

        elif (
            pd.notna(foreign_flow)
            and foreign_flow < 0
        ):
            ws["G12"].fill = PatternFill(
                "solid",
                fgColor=RED,
            )

        else:
            ws["G12"].fill = PatternFill(
                "solid",
                fgColor=YELLOW,
            )

        ws["G12"].font = Font(bold=True)

        self.set_section_title(
            ws,
            15,
            "How to Read This Tab",
            8,
        )

        guide_rows = [
            [
                "Foreign Net Flow Proxy",
                (
                    "Foreign-only broker buy value minus "
                    "foreign-only broker sell value across "
                    "the selected date window."
                ),
            ],
            [
                "Top 3 Accumulating Brokers",
                (
                    "The three brokers with the largest "
                    "positive net buy value across the "
                    "selected window."
                ),
            ],
            [
                "Top 3 Distributing Brokers",
                (
                    "The three brokers with the largest "
                    "negative net buy value across the "
                    "selected window."
                ),
            ],
            [
                "Broker-Flow Proxy",
                (
                    "Broker codes show transaction routing "
                    "activity. They do not prove the final "
                    "beneficial owner is institutional, "
                    "foreign, retail, or a specific party."
                ),
            ],
        ]

        for row_number, row_values in enumerate(
            guide_rows,
            start=16,
        ):
            ws.cell(
                row_number,
                1,
                row_values[0],
            )

            ws.cell(
                row_number,
                2,
                row_values[1],
            )

            ws.cell(
                row_number,
                1,
            ).font = Font(bold=True)

        ws.column_dimensions["A"].width = 24
        ws.column_dimensions["B"].width = 34
        ws.column_dimensions["C"].width = 46
        ws.column_dimensions["D"].width = 46
        ws.column_dimensions["E"].width = 24
        ws.column_dimensions["F"].width = 28
        ws.column_dimensions["G"].width = 32
        ws.column_dimensions["H"].width = 50

        ws.freeze_panes = "A11"

    def build_weekly_monthly(
        self,
        df: pd.DataFrame,
    ) -> None:
        ws = self.workbook.create_sheet(
            "6. Weekly & Monthly"
        )

        self.set_title(
            ws,
            "Higher-Timeframe Swing Structure",
            8,
        )

        indexed = df.set_index("Date").sort_index()

        weekly = pd.DataFrame(
            {
                "Open": indexed["Open"].resample("W-FRI").first(),
                "High": indexed["High"].resample("W-FRI").max(),
                "Low": indexed["Low"].resample("W-FRI").min(),
                "Close": indexed["Close"].resample("W-FRI").last(),
                "Volume": indexed["Volume"].resample("W-FRI").sum(),
                "Foreign Flow": indexed["Net Foreign Flow"].resample("W-FRI").sum(min_count=1),
            }
        ).dropna(subset=["Open"])

        weekly["Return"] = (
            (weekly["Close"] - weekly["Open"])
            / weekly["Open"]
        )

        self.set_section_title(
            ws,
            3,
            "Weekly Aggregates",
            8,
        )

        self.set_header(
            ws,
            4,
            [
                "Week Ending",
                "Open",
                "High",
                "Low",
                "Close",
                "Total Volume",
                "Net Foreign Flow",
                "Weekly Return",
            ],
        )

        for row_number, (date, row) in enumerate(
            weekly.iterrows(),
            start=5,
        ):
            values = [
                date,
                row["Open"],
                row["High"],
                row["Low"],
                row["Close"],
                row["Volume"],
                row["Foreign Flow"],
                row["Return"],
            ]

            for column, value in enumerate(
                values,
                start=1,
            ):
                ws.cell(
                    row_number,
                    column,
                    self.clean_value(value),
                )

        monthly_header_row = ws.max_row + 3

        monthly = pd.DataFrame(
            {
                "Open": indexed["Open"].resample("ME").first(),
                "High": indexed["High"].resample("ME").max(),
                "Low": indexed["Low"].resample("ME").min(),
                "Close": indexed["Close"].resample("ME").last(),
                "Volume": indexed["Volume"].resample("ME").sum(),
                "Foreign Flow": indexed["Net Foreign Flow"].resample("ME").sum(min_count=1),
            }
        ).dropna(subset=["Close"])

        monthly["Return"] = (
            (monthly["Close"] - monthly["Open"])
            / monthly["Open"]
        )

        self.set_section_title(
            ws,
            monthly_header_row,
            "Monthly Aggregates",
            8,
        )

        self.set_header(
            ws,
            monthly_header_row + 1,
            [
                "Month Ending",
                "High",
                "Low",
                "Close",
                "Total Volume",
                "Net Foreign Flow",
                "Monthly Return",
            ],
        )

        for row_number, (date, row) in enumerate(
            monthly.iterrows(),
            start=monthly_header_row + 2,
        ):
            values = [
                date,
                row["High"],
                row["Low"],
                row["Close"],
                row["Volume"],
                row["Foreign Flow"],
                row["Return"],
            ]

            for column, value in enumerate(
                values,
                start=1,
            ):
                ws.cell(
                    row_number,
                    column,
                    self.clean_value(value),
                )

        for row in ws.iter_rows():
            for cell in row:
                if cell.column == 1:
                    cell.number_format = "yyyy-mm-dd"

                elif cell.column in [2, 3, 4, 5]:
                    cell.number_format = "#,##0.00"

                elif cell.column in [6, 7]:
                    cell.number_format = "#,##0"

                elif cell.column == 8:
                    cell.number_format = "0.00%"

        for column in range(1, 9):
            ws.column_dimensions[
                get_column_letter(column)
            ].width = 20

        ws.freeze_panes = "A5"

    def build_indicator_guide(self) -> None:
        ws = self.workbook.create_sheet(
            "7. Indicator Guide"
        )

        self.set_title(
            ws,
            "Indicator Guide: Meaning and Practical Use",
            4,
        )

        self.set_header(
            ws,
            3,
            [
                "Indicator",
                "What It Measures",
                "How This System Interprets It",
                "Important Limitation",
            ],
        )

        rows = [
            [
                "SMA20 / SMA50",
                "Average close over 20 and 50 trading days.",
                "Close above both averages supports bullish swing trend bias.",
                "Moving averages lag price and do not predict reversals.",
            ],
            [
                "EMA9 / EMA21",
                "Faster moving averages that react more quickly to price.",
                "EMA9 above EMA21 supports short-term upward momentum.",
                "Can generate false signals in sideways markets.",
            ],
            [
                "Volume Ratio",
                "Today’s volume divided by 20-day average volume.",
                "Above 1.50x indicates notably high participation.",
                "High volume alone does not indicate whether the move will continue.",
            ],
            [
                "RSI14",
                "Speed and magnitude of recent price changes.",
                "45 to 65 is treated as constructive momentum. Above 70 is more extended.",
                "Strong trends can remain overbought or oversold for extended periods.",
            ],
            [
                "MACD",
                "Difference between fast and slow exponential moving averages.",
                "MACD above its signal line indicates improving momentum.",
                "MACD is a lagging momentum indicator.",
            ],
            [
                "ADX14",
                "Strength of a trend, not its direction.",
                "Above 25 suggests a stronger trend. Below 20 suggests weak or ranging conditions.",
                "Use Plus DI and Minus DI for directional context.",
            ],
            [
                "ATR14",
                "Average true price range over 14 days.",
                "Used to set dynamic stop-loss and target distances.",
                "ATR measures volatility, not directional bias.",
            ],
            [
                "Bollinger Bands",
                "Volatility bands around a 20-day average.",
                "Price near upper/lower bands can show expansion or extension.",
                "Touching a band is not automatically a sell or buy signal.",
            ],
            [
                "OBV",
                "Cumulative volume adjusted for up/down price movement.",
                "Rising OBV can support price accumulation.",
                "OBV can be distorted by unusual single-day volume.",
            ],
            [
                "Relative Strength vs IHSG",
                "Stock performance relative to IHSG benchmark.",
                "Rising 20-day ratio means outperformance against the market.",
                "Benchmark comparison does not replace company-specific analysis.",
            ],
            [
                "Foreign Flow",
                "Net foreign buy/sell value, when local data is supplied.",
                "Positive 14-day cumulative flow supports institutional demand.",
                "Unavailable until a reliable local IDX provider is connected.",
            ],
            [
                "Broker Accumulation",
                "Net buying or selling by top brokers, when supplied.",
                "Positive Top-3 14-day net buying supports accumulation scoring.",
                "Broker-code interpretation needs a verified, current classification source.",
            ],
        ]

        self.write_rows(ws, 4, rows)
        self.zebra_rows(ws, 4, ws.max_row, 4)

        ws.column_dimensions["A"].width = 26
        ws.column_dimensions["B"].width = 44
        ws.column_dimensions["C"].width = 52
        ws.column_dimensions["D"].width = 52

        ws.freeze_panes = "A4"

    def build_raw_data(
        self,
        df: pd.DataFrame,
        metrics: dict[str, Any],
    ) -> None:
        ws = self.workbook.create_sheet(
            "8. Raw Daily Data"
        )

        export_df = df.copy()

        # Constant trade-plan levels are repeated for every
        # row so Excel can display them as horizontal lines.
        export_df["Pullback Stop Loss"] = metrics[
            "pullback_stop_loss"
        ]

        export_df["Pullback Target 1"] = metrics[
            "pullback_target_1"
        ]

        export_df["Breakout Trigger"] = metrics[
            "breakout_entry"
        ]

        export_df["Breakout Target 1"] = metrics[
            "breakout_target_1"
        ]

        export_df["Breakout Target 2"] = metrics[
            "breakout_target_2"
        ]

        self.set_title(
            ws,
            "Raw Daily Market Data and Calculated Indicators",
            39,
        )

        columns = [
            "Date",
            "Open",
            "High",
            "Low",
            "Close",
            "Adj Close",
            "Volume",
            "20D Vol SMA",
            "Volume Ratio",
            "Net Foreign Flow",
            "Top 3 Accumulating Brokers",
            "Top 3 Distributing Brokers",
            "Daily Return",
            "Log Return",
            "SMA20",
            "SMA50",
            "EMA9",
            "EMA21",
            "ATR14",
            "ATR Percent",
            "RSI14",
            "MACD",
            "MACD Signal",
            "MACD Histogram",
            "ADX14",
            "Plus DI",
            "Minus DI",
            "BB Upper",
            "BB Middle",
            "BB Lower",
            "Support 20D",
            "Resistance 20D",
            "IHSG Close",
            "RS vs IHSG 20D Change",
            "Pullback Stop Loss",
            "Pullback Target 1",
            "Breakout Trigger",
            "Breakout Target 1",
            "Breakout Target 2",
        ]

        self.set_header(
            ws,
            3,
            columns,
        )

        for row_number, (_, row) in enumerate(
            export_df.iterrows(),
            start=4,
        ):
            for column_number, column_name in enumerate(
                columns,
                start=1,
            ):
                ws.cell(
                    row_number,
                    column_number,
                    self.clean_value(row[column_name]),
                )

        self.zebra_rows(
            ws,
            4,
            ws.max_row,
            len(columns),
        )

        for row in range(4, ws.max_row + 1):
            ws.cell(row, 1).number_format = "yyyy-mm-dd"

            for column in range(2, 7):
                ws.cell(
                    row,
                    column,
                ).number_format = "#,##0.00"

            ws.cell(row, 7).number_format = "#,##0"
            ws.cell(row, 8).number_format = "#,##0"
            ws.cell(row, 9).number_format = '0.00"x"'
            ws.cell(row, 10).number_format = "#,##0"
            ws.cell(row, 13).number_format = "0.00%"
            ws.cell(row, 14).number_format = "0.0000"
            ws.cell(row, 20).number_format = "0.00%"
            ws.cell(row, 34).number_format = "0.00%"

        ws.auto_filter.ref = (
            f"A3:{get_column_letter(len(columns))}{ws.max_row}"
        )

        ws.conditional_formatting.add(
            f"I4:I{ws.max_row}",
            CellIsRule(
                operator="greaterThan",
                formula=["1.5"],
                fill=PatternFill(
                    "solid",
                    fgColor=GREEN,
                ),
            ),
        )

        ws.conditional_formatting.add(
            f"J4:J{ws.max_row}",
            CellIsRule(
                operator="greaterThan",
                formula=["0"],
                fill=PatternFill(
                    "solid",
                    fgColor=GREEN,
                ),
            ),
        )

        ws.conditional_formatting.add(
            f"J4:J{ws.max_row}",
            CellIsRule(
                operator="lessThan",
                formula=["0"],
                fill=PatternFill(
                    "solid",
                    fgColor=RED,
                ),
            ),
        )

        for column in range(1, len(columns) + 1):
            letter = get_column_letter(column)
            ws.column_dimensions[letter].width = 16

        ws.column_dimensions["A"].width = 14
        ws.column_dimensions["K"].width = 30
        ws.column_dimensions["L"].width = 30

        ws.freeze_panes = "A4"

    def save_report(
        self,
        output_path: Path,
    ) -> None:
        for ws in self.workbook.worksheets:
            for row in ws.iter_rows():
                for cell in row:
                    if (
                        cell.row > 1
                        and cell.value is not None
                    ):
                        cell.alignment = Alignment(
                            vertical="center",
                            wrap_text=(
                                cell.column in [2, 3, 4, 11, 12]
                            ),
                        )

        self.workbook.save(output_path)

        for chart_path in self.temporary_chart_files:
            try:
                if chart_path.exists():
                    chart_path.unlink()

            except OSError as error:
                logger.warning(
                    "Unable to remove temporary chart %s: %s",
                    chart_path,
                    error,
                )

        self.temporary_chart_files.clear()

        logger.info(
            "Saved report: %s",
            output_path,
        )


# ============================================================
# MAIN APPLICATION
# ============================================================

def main(
    preselected_ticker: str | None = None,
    skip_workflow_menu: bool = False,
) -> None:
    if not skip_workflow_menu:
        print("=" * 68)
        print(
            "AUTOMATED IDX SWING TRADING & "
            "MARKET STRUCTURE SYSTEM"
        )
        print("=" * 68)
        print()
        print("Select workflow:")
        print("1. Single Stock Deep-Dive Report")
        print("2. Batch Watchlist Screener")
        print("3. Next-Session & 5-Session Scenario Projector")
        print()

        workflow_choice = input(
            "Select workflow [1/2/3]: "
        ).strip()

        if workflow_choice == "3":
            from run_predictor import main as run_predictor

            run_predictor()
            return

        if workflow_choice == "2":
            from run_screener import main as run_screener

            run_screener()
            return

        if workflow_choice not in {"1", "2", "3"}:
            print()
            print("Invalid workflow selection.")
            print("Please choose 1, 2, or 3.")
            return

        print()

    ticker_input = (
        preselected_ticker.strip().upper()
        if preselected_ticker
        else input(
            "Enter IDX ticker code (example: MDIA): "
        ).strip().upper()
    )

    if preselected_ticker:
        print(
            "Selected screener finalist: "
            f"{ticker_input}"
        )

    if not ticker_input:
        print("Ticker code cannot be empty.")
        return

    print()
    print("=" * 68)
    print("BROKER SUMMARY & FOREIGN FLOW SOURCE")
    print("=" * 68)
    print("1. Yahoo-only analysis")
    print("2. Manual Mirae CSV input")
    print("3. Index Alpha API")
    print()

    data_source_choice = input(
        "Select source [1/2/3]: "
    ).strip()

    source_options = {
        "1": "YAHOO_ONLY",
        "2": "MANUAL_MIRAE",
        "3": "INDEX_ALPHA_API",
    }

    flow_data_source_mode = source_options.get(
        data_source_choice,
        "YAHOO_ONLY",
    )

    broker_lookback_days = 14
    use_api_cache = True

    if flow_data_source_mode == "MANUAL_MIRAE":
        print()
        print(
            "Manual Mirae CSV mode selected."
        )
        print(
            "Expected files:"
        )
        print(
            "data/manual/broker_summary_manual.csv"
        )
        print(
            "data/manual/foreign_flow_manual.csv"
        )

    elif flow_data_source_mode == "INDEX_ALPHA_API":
        print()

        lookback_input = input(
            "Broker-data lookback days [default: 14]: "
        ).strip()

        if lookback_input:
            try:
                broker_lookback_days = int(
                    lookback_input
                )

                if broker_lookback_days < 1:
                    raise ValueError

            except ValueError:
                print(
                    "Invalid lookback. "
                    "Using the default 14 days."
                )

                broker_lookback_days = 14

        cache_input = input(
            "Use cached API data if available? "
            "[Y/N, default: Y]: "
        ).strip().upper()

        use_api_cache = cache_input != "N"

    else:
        print()
        print(
            "Yahoo-only analysis selected. "
            "Broker and foreign-flow score criteria "
            "will be excluded."
        )

    try:
        print()
        print("Downloading market data and calculating indicators...")
        print()

        fetcher = StockDataFetcher()
        analytics = AnalyticsEngine()

        yahoo_data = fetcher.fetch_yahoo_data(
            ticker_input,
            period="2y",
        )

        # ----------------------------------------------------
        # Select the appropriate Broker Summary / Foreign Flow
        # data source based on the user's terminal selection.
        # ----------------------------------------------------

        local_idx_data = pd.DataFrame()

        api_calls_used = 0
        broker_data_source = "Not Connected"

        # ----------------------------------------------------
        # Minervini preflight gate for Index Alpha API usage.
        #
        # This checks the long-term trend before optional API
        # requests are made, preserving limited API quota.
        # ----------------------------------------------------

        api_skipped_by_trend_template = False

        if flow_data_source_mode == "INDEX_ALPHA_API":
            preflight_df = analytics.calculate_indicators(
                yahoo_data.daily_price,
                yahoo_data.benchmark_price,
                pd.DataFrame(),
            )

            preflight_minervini = (
                analytics.evaluate_minervini_template(
                    preflight_df
                )
            )

            if not preflight_minervini["passed"]:
                print()
                print("=" * 68)
                print("MINERVINI TREND TEMPLATE FAILED")
                print("=" * 68)

                print(
                    "Passed checks: "
                    f"{preflight_minervini['passed_checks']}"
                    f"/{preflight_minervini['total_checks']}"
                )

                print()

                for failed_rule in preflight_minervini[
                    "failed_rules"
                ]:
                    print(f"- {failed_rule}")

                print()

                override_input = input(
                    "Use Index Alpha API anyway? [Y/N]: "
                ).strip().upper()

                api_skipped_by_trend_template = (
                    override_input != "Y"
                )

                if api_skipped_by_trend_template:
                    print()
                    print(
                        "Index Alpha API skipped. "
                        "Continuing with Yahoo-only analysis."
                    )

        # ----------------------------------------------------
        # Option 1: Yahoo-only analysis
        # ----------------------------------------------------

        if flow_data_source_mode == "YAHOO_ONLY":
            print()
            print(
                "Running Yahoo-only analysis."
            )

            broker_data_source = (
                "Yahoo Only "
                "(No Broker / Foreign Flow Data)"
            )

        # ----------------------------------------------------
        # Option 2: Manual Mirae CSV input
        # ----------------------------------------------------

        elif flow_data_source_mode == "MANUAL_MIRAE":
            manual_loader = ManualMiraeDataLoader()

            local_idx_data = manual_loader.load(
                yahoo_data.ticker
            )

            if local_idx_data.empty:
                print()
                print(
                    "No usable Manual Mirae data was found "
                    f"for {yahoo_data.ticker}."
                )

                print(
                    "Continuing with Yahoo-only analysis."
                )

                broker_data_source = (
                    "Manual Mirae CSV "
                    "(No Matching Data Found)"
                )
            else:
                print()
                print(
                    "Manual Mirae data loaded successfully."
                )

                print(
                    f"Manual records loaded: "
                    f"{len(local_idx_data)}"
                )

                broker_data_source = (
                    "Manual Mirae CSV Input"
                )

        # ----------------------------------------------------
        # Option 3: Index Alpha API
        # ----------------------------------------------------

        elif (
            flow_data_source_mode == "INDEX_ALPHA_API"
            and not api_skipped_by_trend_template
        ):
            broker_fetcher = IndexAlphaBrokerFetcher()

            if not broker_fetcher.is_configured():
                print()
                print(
                    "Index Alpha API was selected, but no "
                    "real API key was found in .env."
                )

                print(
                    "Continuing with Yahoo-only analysis."
                )

                broker_data_source = (
                    "Index Alpha API "
                    "(API Key Unavailable)"
                )

            else:
                try:
                    api_broker_data, _, api_calls_used = (
                        broker_fetcher.fetch_aggregated_window(
                            ticker=yahoo_data.ticker.replace(
                                ".JK",
                                "",
                            ),
                            price_dates=yahoo_data.daily_price[
                                "Date"
                            ],
                            lookback_days=broker_lookback_days,
                            use_cache=use_api_cache,
                        )
                    )

                    if api_broker_data.empty:
                        print()
                        print(
                            "Index Alpha returned no usable "
                            "broker data for this ticker/date range."
                        )

                        print(
                            "Continuing with Yahoo-only analysis."
                        )

                        broker_data_source = (
                            "Index Alpha API "
                            "(No Data Returned)"
                        )

                    else:
                        local_idx_data = api_broker_data

                        broker_data_source = (
                            "Index Alpha API "
                            f"({broker_lookback_days}-day "
                            "aggregated window)"
                        )

                        print()
                        print(
                            "Index Alpha broker data "
                            "loaded successfully."
                        )

                        print(
                            f"API calls used in this run: "
                            f"{api_calls_used}"
                        )

                except Exception as api_error:
                    logger.warning(
                        "Index Alpha integration failed: %s",
                        api_error,
                    )

                    print()
                    print(
                        "Index Alpha request failed. "
                        "Continuing with Yahoo-only analysis."
                    )

                    print(
                        f"API error: {api_error}"
                    )

                    broker_data_source = (
                        "Index Alpha API "
                        "(Request Failed)"
                    )

        # ----------------------------------------------------
        # Index Alpha skipped because Minervini failed and the
        # user did not authorize an override.
        # ----------------------------------------------------

        elif (
            flow_data_source_mode == "INDEX_ALPHA_API"
            and api_skipped_by_trend_template
        ):
            broker_data_source = (
                "Index Alpha API "
                "(Skipped: Trend Template Failed)"
            )

        analyzed_df = analytics.calculate_indicators(
            yahoo_data.daily_price,
            yahoo_data.benchmark_price,
            local_idx_data,
        )

        # ----------------------------------------------------
        # Keep 2 years internally for SMA150, SMA200,
        # Minervini logic, and 52-week calculations.
        #
        # Export only the most recent six months to Excel.
        # ----------------------------------------------------

        latest_market_date = analyzed_df["Date"].max()

        six_month_cutoff = (
            latest_market_date
            - pd.DateOffset(months=6)
        )

        report_df = analyzed_df[
            analyzed_df["Date"] >= six_month_cutoff
        ].copy()

        logger.info(
            "Internal analysis records: %s | "
            "Excel reporting records: %s",
            len(analyzed_df),
            len(report_df),
        )

        fundamentals = analytics.extract_fundamentals(
            yahoo_data.info,
            yahoo_data.balance_sheet,
            yahoo_data.income_statement,
            yahoo_data.cash_flow,
        )

        metrics = analytics.calculate_metrics(
            analyzed_df,
            fundamentals,
            report_df,
        )

        generated_datetime = datetime.now()

        generated_at_display = generated_datetime.strftime(
            "%Y-%m-%d %H:%M"
        )

        generated_at_filename = generated_datetime.strftime(
            "%Y%m%d_%H%M"
        )

        ticker_code = yahoo_data.ticker.replace(
            ".JK",
            "",
        )

        output_file = (
            OUTPUT_DIR
            / (
                f"{ticker_code}_Swing_Report_"
                f"{generated_at_filename}.xlsx"
            )
        )

        report = ExcelReportBuilder()

        report.build_dashboard(
            yahoo_data.ticker,
            fundamentals,
            metrics,
            generated_at_display,
            broker_data_source,
            api_calls_used,
            (
                broker_lookback_days
                if flow_data_source_mode == "INDEX_ALPHA_API"
                else None
            ),
            report_df,
        )

        report.build_trade_plan(
            yahoo_data.ticker,
            metrics,
        )

        report.build_raw_data(
            report_df,
            metrics,
        )

        report.build_technicals(
            yahoo_data.ticker,
            metrics,
        )

        report.build_fundamentals(
            fundamentals
        )

        report.build_bandarmology(
            report_df,
            broker_data_source,
            (
                broker_lookback_days
                if flow_data_source_mode == "INDEX_ALPHA_API"
                else None
            ),
        )

        report.build_weekly_monthly(
            report_df
        )

        report.build_indicator_guide()

        report.save_report(output_file)

        pdf_output_file = output_file.with_suffix(".pdf")

        generate_pdf_report(
            output_path=pdf_output_file,
            ticker=yahoo_data.ticker,
            company_name=fundamentals["name"],
            latest_close=metrics["latest_close"],
            metrics=metrics,
        )

        logger.info(
            "Saved PDF summary: %s",
            pdf_output_file,
        )

        print("=" * 68)
        print("REPORT GENERATED SUCCESSFULLY")
        print("=" * 68)
        print(f"Ticker: {yahoo_data.ticker}")
        print(f"Company: {fundamentals['name']}")
        print(f"Latest Close: {metrics['latest_close']:,.2f}")
        print(f"Decision: {metrics['decision']}")
        print(
            "Raw Score: "
            f"{metrics['raw_score']} / "
            f"{metrics['eligible_max_score']} eligible points"
        )

        print(
            "Normalized Score: "
            f"{metrics['normalized_score']:.1f} / 100"
        )

        print(
            "Data Coverage: "
            f"{metrics['data_coverage_percent']:.0f}%"
        )

        print(
            "Broker Summary Data: "
            + (
                "Connected"
                if metrics["broker_data_available"]
                else "Not Connected"
            )
        )

        print(
            "Foreign Flow Data: "
            + (
                "Connected"
                if metrics["foreign_flow_data_available"]
                else "Not Connected"
            )
        )
        print(f"Wyckoff Phase: {metrics['wyckoff_phase']}")
        print(
            "Minervini Trend Template: "
            + (
                "PASSED"
                if metrics["minervini_passed"]
                else "FAILED"
            )
            + (
                f" ({metrics['minervini_passed_checks']}"
                f"/{metrics['minervini_total_checks']} checks)"
            )
        )
        print(
            "Extension Risk: "
            f"{metrics['extension_risk_status']}"
        )
        print(f"Risk: {metrics['risk_label']}")
        print()
        print("Generated Excel file:")
        print(output_file)

        print()
        print("Generated PDF summary:")
        print(pdf_output_file)

    except Exception as error:
        logger.exception("Application failed")

        print()
        print("=" * 68)
        print("APPLICATION ERROR")
        print("=" * 68)
        print(error)
        print()
        print(
            "Open logs/stock_analysis.log "
            "for full technical details."
        )


if __name__ == "__main__":
    main()